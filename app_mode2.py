from datetime import date
from hashlib import sha256
from html import escape
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import pandas as pd
import streamlit as st
import excel_reports as reports
from official_template import load_official_halfyear_template
from report_info import infer_report_info, apply_report_info, report_filename
from mode2_style import MODE2_CSS

reports._load_halfyear_template = load_official_halfyear_template


def _restore(info):
    for key, field in [('final_report_year', 'year'), ('final_report_label', 'half'),
                       ('final_period_start', 'start'), ('final_period_end', 'end')]:
        st.session_state[key] = info[field]
    st.session_state.final_report_selection = {field: info[field] for field in ("year", "half", "start", "end")}


def _heading(number, title, description, color='blue'):
    st.markdown(f'<div class="final-step-heading {color}"><span>{number}</span><b>{title}</b></div><div class="final-step-desc">{description}</div>', unsafe_allow_html=True)


@st.cache_data(show_spinner=False)
def _label_report(content, year, half, start, end):
    return apply_report_info(content, year, half, start, end)


@st.cache_data(show_spinner=False, max_entries=8)
def _sheet_preview(content, sheet_index):
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=False)
    try:
        sheet = workbook.worksheets[sheet_index]
        rows = list(sheet.iter_rows(values_only=True))
        while rows and all(value is None for value in rows[-1]):
            rows.pop()
        width = max((i + 1 for row in rows for i, value in enumerate(row) if value is not None), default=0)
        # Strings preserve mixed Excel cell types without Arrow conversion failures.
        frame = pd.DataFrame([[str(value) if value is not None else '' for value in row[:width]] for row in rows],
                             columns=[get_column_letter(i + 1) for i in range(width)],
                             index=range(1, len(rows) + 1))
        frame.index.name = '행'
        return sheet.title, frame
    finally:
        workbook.close()


def render_mode2():
    st.markdown(MODE2_CSS, unsafe_allow_html=True)
    left, right = st.columns([1.15, 1], gap='small')
    with left:
        with st.container(border=True, key='final_upload_card'):
            st.markdown('<div class="panel-head"><div class="panel-icon">▤</div><div><div class="panel-title">자료 입력</div><div class="panel-desc">검토 완료된 Excel 기초자료 파일을 업로드하세요.</div></div></div>', unsafe_allow_html=True)
            upload = st.file_uploader('검토 완료 1-4 기초자료', type=['xlsx'], accept_multiple_files=False,
                                      key='review_upload', label_visibility='collapsed')
            st.markdown('<div class="upload-checks"><span>✓ 자료 집계 화면에서 내려받은 검토용 기초자료를 사용합니다.</span><span>✓ 사용자 수정값을 우선 반영하여 최종 4시트 보고서를 만듭니다.</span><span>✓ 업로드한 파일은 현재 작업 중에만 임시 사용됩니다.</span></div>', unsafe_allow_html=True)
    result = None
    error = None
    if upload:
        fingerprint = sha256(upload.getvalue()).hexdigest()
        if st.session_state.get('final_source_fingerprint') != fingerprint:
            try:
                loaded = reports.build_final_halfyear_report_bytes(upload.getvalue())
                info = infer_report_info(loaded[1], loaded[2].get('title', ''))
                st.session_state.final_source_result = loaded
                st.session_state.final_source_info = info
                st.session_state.final_source_fingerprint = fingerprint
                _restore(info)
                st.session_state.final_show_review = False
            except Exception as exc:
                error = str(exc)
        if not error:
            result = st.session_state.final_source_result
    fallback = dict(year=date.today().year, half='상반기', start=date(date.today().year, 1, 1),
                    end=date(date.today().year, 12, 31), warnings=[], first=0, second=0, source='파일 업로드 후 자동 인식')
    info = st.session_state.final_source_info if result else fallback
    for key, field in [('final_report_year', 'year'), ('final_report_label', 'half'),
                       ('final_period_start', 'start'), ('final_period_end', 'end')]:
        st.session_state.setdefault(key, st.session_state.get('final_report_selection', info)[field])
    meta = result[2] if result else {}
    institution = meta.get('institution', '') or '기관명미입력'
    region = meta.get('region', '') or '미인식'
    with right:
        with st.container(border=True, key='final_conditions'):
            st.markdown('<div class="final-condition-title"><svg viewBox="0 0 24 24" width="24" height="24" fill="none" stroke="currentColor" stroke-width="2.5" aria-hidden="true"><path d="M3 4h18l-7 8v8l-4-2v-6z"/></svg> 생성 조건 <span>자동 인식 후 수정할 수 있습니다.</span></div>', unsafe_allow_html=True)
            a, b = st.columns(2)
            year = int(a.number_input('보고연도', min_value=1900, max_value=2100, step=1,
                                      key='final_report_year', disabled=not result))
            half = b.selectbox('보고구분', ['상반기', '하반기'], key='final_report_label', disabled=not result)
            a, b = st.columns(2)
            start = a.date_input('기간 시작', key='final_period_start', disabled=not result,
                                 min_value=date(1900, 1, 1), max_value=date(2100, 12, 31))
            end = b.date_input('기간 종료', key='final_period_end', disabled=not result,
                               min_value=date(1900, 1, 1), max_value=date(2100, 12, 31))
            filename = report_filename(year, half, institution)
            st.markdown(f'<div class="final-recognized"><span>기관명 <b>{escape(institution) if result else "업로드 후 자동 인식"}</b></span><span>기준 지역 <b>{escape(region) if result else "—"}</b></span></div><div class="final-filename"><small>보고연도 + 보고구분 + 기관명 → 파일명</small><b>{escape(filename) if result else "검토파일을 업로드하면 파일명이 표시됩니다."}</b></div>', unsafe_allow_html=True)
            with st.expander('자동 인식 근거 · 보고서 반영 정보'):
                if result:
                    st.write(f'추천 기준: {info["year"]}년 1~6월 {info["first"]:,}건 / 7~12월 {info["second"]:,}건')
                    st.caption(f'기간 인식: {info["source"]}')
                    for warning in info['warnings']:
                        st.warning(warning)
                    st.write(f'제목: {year}년 {half} 지역경제활성화 추진 실적')
                    st.write(f'보고기간: {start:%Y.%m.%d} ~ {end:%Y.%m.%d}')
                    st.caption('기간은 보고서 표기입니다. 업로드한 계약을 제외하거나 금액을 재분류하지 않습니다.')
                    st.button('자동 인식값으로 되돌리기', key='final_restore', on_click=_restore, args=(info,))
                else:
                    st.caption('계약일 건수로 연도·반기를 추천하고, 파일의 집계기간을 우선 인식합니다.')
    if result:
        st.session_state.final_report_selection = dict(year=year, half=half, start=start, end=end)
    if error:
        st.error(f'검토파일을 처리하지 못했습니다: {error}')
    valid = bool(result) and start <= end
    if result and start > end:
        st.error('기간 종료는 기간 시작보다 빠를 수 없습니다.')
    if result and info['warnings']:
        st.warning('날짜 자동 인식에 참고할 사항이 있습니다. 생성 조건의 인식 근거를 확인하고 필요하면 수정하세요. 별도 확인 체크 없이 다운로드할 수 있습니다.')
    records = result[1] if result else []
    notes = []
    for key, label in [('blank_purpose', '구입목적 미입력 → 그 외'), ('invalid_purpose', '구입목적 보정 → 그 외'), ('corrected_location', '소재지 보정')]:
        if meta.get(key):
            notes.append(f'{label} {meta[key]:,}건')
    correction_count = sum(meta.get(k, 0) for k in ('blank_purpose', 'invalid_purpose', 'corrected_location'))
    final_bytes = _label_report(result[0], year, half, start, end) if valid else b''
    with st.container(border=True, key='final_steps'):
        st.markdown(f'<div class="final-stage-header"><div><div class="panel-title"><span class="final-pin"></span>최종 보고서 확인 및 다운로드</div><div class="panel-desc">업로드하면 자동 생성됩니다. 보고 정보를 확인하거나 수정한 뒤 바로 다운로드하세요.</div></div><div class="final-status {"attention" if correction_count or error else ""}"><b>{"분류 보정 " + str(correction_count) + "항목" if correction_count else "파일 확인 완료" if result else "검토파일 업로드 대기"}</b><small>{"같은 계약의 보정 항목이 중복 집계될 수 있습니다." if correction_count else "검토한 기초자료를 기준으로 최종 보고서를 작성합니다."}</small></div></div>', unsafe_allow_html=True)
        c1, c2, c3 = st.columns(3, gap='small')
        with c1:
            with st.container(border=True, key='final_check_step'):
                _heading(1, '기초자료 확인 (선택)', '기초자료 구조와 사용자 수정값,<br>기관정보를 확인합니다.')
                if st.button('기초자료 펼치기 / 접기', key='final_check', width='stretch', disabled=not result):
                    st.session_state.final_show_review = not st.session_state.get('final_show_review', False)
                st.markdown(f'<div class="final-count blue"><b>검토 기초자료 {len(records):,}건</b><small>필수 열과 분류값 확인</small></div>', unsafe_allow_html=True)
        with c2:
            with st.container(border=True, key='final_generate_step'):
                _heading(2, '반기보고서 자동 생성', '공사·용역·물품 집계와 수정값을 반영합니다.<br>보고 정보를 바꾸면 결과도 자동 갱신됩니다.', 'green')
                st.markdown(f'<div class="final-auto-status">{"✓ 자동 생성 완료" if valid else "보고기간을 수정해 주세요" if result else "파일 업로드 대기"}</div>', unsafe_allow_html=True)
                st.markdown(f'<div class="final-count green"><b>{"생성 완료 4개 시트" if valid else "생성 양식 4개 시트"}</b><small>공사 / 용역 / 물품 / 검토반영</small></div>', unsafe_allow_html=True)
        with c3:
            with st.container(border=True, key='final_download_step'):
                _heading(3, '최종 파일 다운로드', '보고 정보와 파일명을 확인하고<br>최종 파일을 내려받으세요.', 'orange')
                st.download_button('최종 반기보고서 다운로드', final_bytes, filename,
                                   'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                   key='final_halfyear_download', width='stretch', disabled=not valid)
                st.markdown(f'<div class="final-count orange"><b>다운로드 가능 {1 if valid else 0}개 파일</b><small>수정한 보고 정보 · 파일명 반영</small></div>', unsafe_allow_html=True)
        st.markdown('<div class="final-workflow"><b>ⓘ 작업 순서 안내</b><span>❶ 파일 업로드 → ❷ 보고 정보 확인·수정 → ❸ 바로 다운로드</span><details><summary>사용방법 보기</summary>검토용 기초자료 1개를 업로드하고 자동 인식된 보고 정보를 확인하세요. 수정한 제목·기간·파일명은 즉시 보고서에 반영됩니다. 별도 확인 체크나 생성 버튼 없이 다운로드할 수 있습니다. 아래 시트를 선택하면 실제 다운로드할 내용을 미리 볼 수 있습니다.</details></div>', unsafe_allow_html=True)
    if notes:
        st.warning(' · '.join(notes))
    with st.container(border=True, key='final_sheet_preview'):
        st.markdown('### 시트 내용 미리보기')
        if valid:
            selected = st.radio('확인할 시트', ['공사', '용역', '물품', '검토반영'],
                                horizontal=True, key='final_preview_sheet')
            sheet_name, frame = _sheet_preview(final_bytes, ['공사', '용역', '물품', '검토반영'].index(selected))
            st.caption(f'{sheet_name} · 다운로드 파일의 실제 셀 내용입니다. 서식과 병합은 Excel 파일에서 확인하세요.')
            st.dataframe(frame, width='stretch', height=380)
        else:
            st.info('검토파일을 업로드하고 올바른 보고기간을 설정하면 시트 내용을 확인할 수 있습니다.')
    if result:
        base, edu, book = result[3]
        totals = {kind: sum(base[kind, loc][0] for loc in (1, 2, 3)) for kind in ('공사', '용역', '물품')}
        st.markdown(f'<div class="final-result-summary"><b>검토 반영 결과</b><span>기초자료 {len(records):,}건</span><span>공사 {totals["공사"]:,} · 용역 {totals["용역"]:,} · 물품 {totals["물품"]:,}</span><span>교육용 {sum(edu[loc][0] for loc in (1, 2, 3)):,} · 도서 {sum(book[loc][0] for loc in (1, 2, 3)):,}</span></div>', unsafe_allow_html=True)
        if st.session_state.get('final_show_review'):
            st.dataframe(pd.DataFrame(records), hide_index=True, width='stretch')
            st.caption('검토파일의 사용자 수정값을 반영한 결과입니다. 원본 변경은 검토용 Excel에서 진행하세요.')
