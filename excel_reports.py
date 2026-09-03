import ast
import base64
import re
from copy import copy
from functools import lru_cache
from datetime import datetime
from io import BytesIO

import openpyxl
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

from core_logic import (
    CONTRACT_METHODS,
    COMPETITION_METHODS,
    PURCHASE_PURPOSES,
    aggregate_records,
    format_money,
    records_from_source,
)

from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent
HALFYEAR_TEMPLATE = BASE_DIR / "halfy_report_template.xlsx"
LEGACY_APP = BASE_DIR / "legacy_quarter_app.py"
REVIEW_SHEET = "1-4. 기초자료(지역경제활성화)"
SUMMARY_SHEETS = ["1-1. 총괄(공사)", "1-2. 총괄(용역)", "1-3. 총괄(물품)"]


@lru_cache(maxsize=1)
def _load_quarter_template_bytes():
    """기존 앱 소스에 들어 있던 분기보고서 템플릿을 그대로 재사용합니다."""
    source = LEGACY_APP.read_text(encoding="utf-8")
    tree = ast.parse(source)
    for node in tree.body:
        if isinstance(node, (ast.Assign, ast.AnnAssign)):
            targets = node.targets if isinstance(node, ast.Assign) else [node.target]
            if any(isinstance(t, ast.Name) and t.id == "TEMPLATE_BASE64" for t in targets):
                value = ast.literal_eval(node.value)
                return base64.b64decode(value)
    raise RuntimeError("기존 분기보고서 템플릿을 찾을 수 없습니다.")


def build_quarter_report_bytes(results, target_region):
    """기존 분기보고서 서식을 변경하지 않고 집계값만 채웁니다."""
    wb = load_workbook(BytesIO(_load_quarter_template_bytes()))
    ws = wb.active

    for row in ws.iter_rows(min_row=1, max_row=4):
        for cell in row:
            if type(cell).__name__ != "MergedCell" and isinstance(cell.value, str) and "천안" in cell.value:
                cell.value = cell.value.replace("천안", target_region)

    ws["B5"] = results[("공사", 1)][0]; ws["B6"] = results[("공사", 1)][1]
    ws["C5"] = results[("공사", 2)][0]; ws["C6"] = results[("공사", 2)][1]
    ws["D5"] = results[("공사", 3)][0]; ws["D6"] = results[("공사", 3)][1]
    ws["F5"] = results[("용역", 1)][0]; ws["F6"] = results[("용역", 1)][1]
    ws["G5"] = results[("용역", 2)][0]; ws["G6"] = results[("용역", 2)][1]
    ws["H5"] = results[("용역", 3)][0]; ws["H6"] = results[("용역", 3)][1]
    ws["J5"] = results[("물품", 1)][0]; ws["J6"] = results[("물품", 1)][1]
    ws["K5"] = results[("물품", 2)][0]; ws["K6"] = results[("물품", 2)][1]
    ws["L5"] = results[("물품", 3)][0]; ws["L6"] = results[("물품", 3)][1]

    out = BytesIO(); wb.save(out); return out.getvalue()


def _load_halfyear_template():
    if not HALFYEAR_TEMPLATE.exists():
        raise FileNotFoundError("반기보고서 공식 템플릿 파일을 찾을 수 없습니다.")
    return load_workbook(HALFYEAR_TEMPLATE)


def _find_last_data_row(ws, start_row=7, scan_limit=5000):
    last = start_row - 1
    limit = min(ws.max_row, scan_limit)
    blank_run = 0
    for r in range(start_row, limit + 1):
        has_value = any(ws.cell(r, c).value not in (None, "") for c in range(1, min(ws.max_column, 15) + 1))
        if has_value:
            last = r
            blank_run = 0
        else:
            blank_run += 1
            if last >= start_row and blank_run >= 100:
                break
    return last


def _clear_data_values(ws, end_row):
    for r in range(7, max(7, end_row) + 1):
        for c in range(1, 16):
            ws.cell(r, c).value = None


def _reset_review_validations(ws, end_row):
    # 원본 양식의 과거/중복 validation을 정리하고 현재 안내값으로 통일합니다.
    ws.data_validations.dataValidation = []
    end_row = max(end_row, 1000)
    validations = [
        ("D", ["학교(유)", "학교(초)", "학교(중)", "학교(고)", "학교(특수)", "교육지원청", "직속기관"]),
        ("E", ["공사", "용역", "물품"]),
        ("F", CONTRACT_METHODS),
        ("G", COMPETITION_METHODS),
        ("M", ["충남도내(관내)", "충남도내(관외)", "타시도"]),
        ("N", PURCHASE_PURPOSES),
    ]
    for col, items in validations:
        dv = DataValidation(type="list", formula1='"' + ",".join(items) + '"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col}7:{col}{end_row}")


def _amount_label(target_amount):
    if target_amount >= 10000 and target_amount % 10000 == 0:
        return f"{target_amount // 10000:,}만원"
    return f"{target_amount:,}원"


def _review_title(report_year, report_label, target_amount, start_date, end_date):
    return (
        f"붙임1-4. {report_year}년 {report_label} 지역경제활성화 추진 실적"
        f"[기준(건당 {_amount_label(target_amount)} 이상), "
        f"{start_date.year}. {start_date.month}. {start_date.day}. ~ "
        f"{end_date.year}. {end_date.month}. {end_date.day}.]"
    )


def _apply_template_row_style(ws, row, template_row=7):
    if row == template_row:
        return
    ws.row_dimensions[row].height = ws.row_dimensions[template_row].height
    for c in range(1, 16):
        src = ws.cell(template_row, c)
        dst = ws.cell(row, c)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        dst.alignment = copy(src.alignment)


def _write_review_records(ws, records, target_region, institution_name, school_level):
    for i, rec in enumerate(records, start=1):
        r = i + 6
        _apply_template_row_style(ws, r)
        values = [
            i,
            target_region,
            institution_name,
            school_level,
            rec.get("목적물", ""),
            rec.get("계약방법", ""),
            rec.get("견적경쟁방법", ""),
            rec.get("계약명", ""),
            rec.get("계약일자", ""),
            format_money(rec.get("계약금액", 0)),
            rec.get("업체명", ""),
            rec.get("주소", ""),
            rec.get("소재지", ""),
            rec.get("구입목적", ""),
            rec.get("비고", ""),
        ]
        for c, value in enumerate(values, start=1):
            ws.cell(r, c).value = value
        ws.cell(r, 10).number_format = "#,##0"
        if isinstance(ws.cell(r, 9).value, datetime):
            ws.cell(r, 9).number_format = "yyyy-mm-dd"


def build_review_workbook_bytes(
    df,
    headers,
    target_amount,
    target_region,
    institution_name,
    school_level,
    report_year,
    report_label,
    start_date,
    end_date,
):
    """공식 1-4 시트를 그대로 활용한 사용자 검토용 파일을 생성합니다."""
    records = records_from_source(df, headers, target_amount, target_region)
    wb = _load_halfyear_template()
    for name in list(wb.sheetnames):
        if name != REVIEW_SHEET:
            wb.remove(wb[name])
    ws = wb[REVIEW_SHEET]

    old_last = _find_last_data_row(ws)
    _clear_data_values(ws, max(old_last, len(records) + 6))
    ws["A1"] = _review_title(report_year, report_label, target_amount, start_date, end_date)
    ws["A4"] = "◑ 자료관리목록 업로드 → 주소 보완 → 검토용 기초자료 생성 → 사용자 검토 → 최종 반기보고서 생성"
    ws["J4"] = "지역경제활성화 자동 집계 시스템 · Streamlit"
    _write_review_records(ws, records, target_region, institution_name, school_level)
    _reset_review_validations(ws, len(records) + 6)
    ws.auto_filter.ref = f"A6:O{max(7, len(records) + 6)}"
    ws.freeze_panes = "A7"

    out = BytesIO(); wb.save(out)
    return out.getvalue(), len(records), records


def _find_header_row(ws):
    for r in range(1, min(ws.max_row, 30) + 1):
        vals = [str(ws.cell(r, c).value or "").replace("\n", "").strip() for c in range(1, min(ws.max_column, 20) + 1)]
        if "순" in vals and any("계약금액" in v for v in vals) and "소재지" in vals:
            return r
    return None


def parse_review_workbook(file_bytes):
    """검토 완료 파일의 사용자가 수정한 값을 최종값으로 읽습니다."""
    wb = load_workbook(BytesIO(file_bytes))
    source_ws = next((wb[n] for n in wb.sheetnames if "1-4" in n or "기초자료" in n), wb.active)
    hr = _find_header_row(source_ws)
    if hr is None:
        raise ValueError("1-4 기초자료의 열 제목 행을 찾을 수 없습니다.")

    header_map = {}
    for c in range(1, source_ws.max_column + 1):
        text = str(source_ws.cell(hr, c).value or "").replace("\n", "").strip()
        if text:
            header_map[text] = c

    def col_contains(text):
        return next((c for h, c in header_map.items() if text in h), None)

    col = {
        "순": col_contains("순"), "지역": col_contains("지역"), "기관명": col_contains("기관명"), "급별": col_contains("급별"),
        "목적물": col_contains("목적물"), "계약방법": col_contains("계약방법"),
        "견적경쟁방법": col_contains("견적/경쟁방법") or col_contains("견적"),
        "계약명": col_contains("계약명"), "계약일자": col_contains("계약일자"), "계약금액": col_contains("계약금액"),
        "업체명": col_contains("업체명"), "주소": col_contains("주소"), "소재지": col_contains("소재지"),
        "구입목적": col_contains("구입목적"), "비고": col_contains("비고"),
    }
    required = ["목적물", "계약금액", "업체명", "소재지", "구입목적"]
    missing = [k for k in required if not col[k]]
    if missing:
        raise ValueError("필수 열이 없습니다: " + ", ".join(missing))

    from core_logic import categorize_region_code, normalize_contract_type, region_label

    records = []
    blank_purpose = invalid_purpose = corrected_location = 0
    first_region = first_inst = first_level = ""
    last_data = _find_last_data_row(source_ws, hr + 1)

    for r in range(hr + 1, max(hr + 1, last_data) + 1):
        ctype = normalize_contract_type(source_ws.cell(r, col["목적물"]).value)
        amount = format_money(source_ws.cell(r, col["계약금액"]).value)
        contract_name = str(source_ws.cell(r, col["계약명"]).value or "").strip() if col["계약명"] else ""
        company = str(source_ws.cell(r, col["업체명"]).value or "").strip()
        if not ctype and not amount and not contract_name and not company:
            continue
        if ctype not in ["공사", "용역", "물품"]:
            continue

        region = str(source_ws.cell(r, col["지역"]).value or "").strip() if col["지역"] else ""
        inst = str(source_ws.cell(r, col["기관명"]).value or "").strip() if col["기관명"] else ""
        level = str(source_ws.cell(r, col["급별"]).value or "").strip() if col["급별"] else ""
        if not first_region and region: first_region = region
        if not first_inst and inst: first_inst = inst
        if not first_level and level: first_level = level

        address = str(source_ws.cell(r, col["주소"]).value or "").strip() if col["주소"] else ""
        loc = str(source_ws.cell(r, col["소재지"]).value or "").strip()
        if loc not in ["충남도내(관내)", "충남도내(관외)", "타시도"]:
            loc = region_label(categorize_region_code(address, region or first_region))
            corrected_location += 1

        purchase = str(source_ws.cell(r, col["구입목적"]).value or "").strip()
        if ctype == "물품":
            if not purchase:
                purchase = "그 외"; blank_purpose += 1
            elif purchase.startswith("도서"):
                purchase = "도서(간행물 등)"
            elif purchase not in PURCHASE_PURPOSES:
                purchase = "그 외"; invalid_purpose += 1
        else:
            purchase = ""

        records.append({
            "순": source_ws.cell(r, col["순"]).value if col["순"] else len(records) + 1,
            "지역": region, "기관명": inst, "급별": level, "목적물": ctype,
            "계약방법": str(source_ws.cell(r, col["계약방법"]).value or "").strip() if col["계약방법"] else "",
            "견적경쟁방법": str(source_ws.cell(r, col["견적경쟁방법"]).value or "").strip() if col["견적경쟁방법"] else "",
            "계약명": contract_name,
            "계약일자": source_ws.cell(r, col["계약일자"]).value if col["계약일자"] else "",
            "계약금액": amount, "업체명": company, "주소": address, "소재지": loc,
            "구입목적": purchase,
            "비고": str(source_ws.cell(r, col["비고"]).value or "").strip() if col["비고"] else "",
        })

    meta = {
        "source_ws": source_ws,
        "header_row": hr,
        "title": str(source_ws["A1"].value or ""),
        "region": first_region,
        "institution": first_inst,
        "school_level": first_level,
        "blank_purpose": blank_purpose,
        "invalid_purpose": invalid_purpose,
        "corrected_location": corrected_location,
    }
    return records, meta


def _copy_top_values(source_ws, target_ws, rows=6, cols=15):
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            # 병합셀의 좌상단 이외 셀에는 쓰지 않습니다.
            if type(target_ws.cell(r, c)).__name__ != "MergedCell":
                target_ws.cell(r, c).value = source_ws.cell(r, c).value


def _write_final_review_sheet(ws, records, meta):
    old_last = _find_last_data_row(ws)
    _clear_data_values(ws, max(old_last, len(records) + 6))
    source_ws = meta.get("source_ws")
    if source_ws is not None:
        _copy_top_values(source_ws, ws)
    for i, rec in enumerate(records, start=1):
        r = i + 6
        _apply_template_row_style(ws, r)
        values = [
            i, rec.get("지역", ""), rec.get("기관명", ""), rec.get("급별", ""), rec.get("목적물", ""),
            rec.get("계약방법", ""), rec.get("견적경쟁방법", ""), rec.get("계약명", ""), rec.get("계약일자", ""),
            format_money(rec.get("계약금액", 0)), rec.get("업체명", ""), rec.get("주소", ""), rec.get("소재지", ""),
            rec.get("구입목적", ""), rec.get("비고", ""),
        ]
        for c, value in enumerate(values, start=1):
            ws.cell(r, c).value = value
        ws.cell(r, 10).number_format = "#,##0"
    _reset_review_validations(ws, len(records) + 6)
    ws.auto_filter.ref = f"A6:O{max(7, len(records) + 6)}"


def _parse_year_label(title):
    m = re.search(r"(\d{4})년\s*([^\s]+)", title or "")
    return (m.group(1), m.group(2)) if m else (str(datetime.now().year), "반기")


def _fill_pair_group(ws, row, start_col, values):
    total_count = total_amount = 0
    c = start_col
    for loc in (1, 2, 3):
        cnt, amt = values[loc]
        ws.cell(row, c).value = cnt
        ws.cell(row, c + 1).value = format_money(amt)
        total_count += cnt; total_amount += amt; c += 2
    ws.cell(row, c).value = total_count
    ws.cell(row, c + 1).value = format_money(total_amount)


def _fill_summary_sheet(ws, purpose, values, region, institution, school_level, title):
    ws["A1"] = title
    ws["A5"] = 1; ws["B5"] = region; ws["C5"] = institution; ws["D5"] = school_level; ws["E5"] = purpose
    _fill_pair_group(ws, 5, 6, values)


def build_final_halfyear_report_bytes(review_bytes):
    records, meta = parse_review_workbook(review_bytes)
    if not records:
        raise ValueError("집계할 기초자료가 없습니다.")
    base, edu, book = aggregate_records(records)
    wb = _load_halfyear_template()
    year, label = _parse_year_label(meta.get("title", ""))
    region = meta.get("region", "")
    institution = meta.get("institution", "")
    school_level = meta.get("school_level", "")

    _fill_summary_sheet(
        wb["1-1. 총괄(공사)"], "공사", {loc: base[("공사", loc)] for loc in (1, 2, 3)},
        region, institution, school_level,
        f"붙임 1-1. {year}년 {label} 지역경제활성화 추진 실적(총괄+공사)",
    )
    _fill_summary_sheet(
        wb["1-2. 총괄(용역)"], "용역", {loc: base[("용역", loc)] for loc in (1, 2, 3)},
        region, institution, school_level,
        f"붙임 1-2. {year}년 {label} 지역경제활성화 추진 실적(총괄+용역)",
    )

    ws3 = wb["1-3. 총괄(물품)"]
    ws3["A1"] = f"붙임 1-3. {year}년 {label} 지역경제활성화 추진 실적(총괄+물품)"
    ws3["A6"] = 1; ws3["B6"] = region; ws3["C6"] = institution; ws3["D6"] = school_level; ws3["E6"] = "물품"
    _fill_pair_group(ws3, 6, 6, {loc: base[("물품", loc)] for loc in (1, 2, 3)})
    _fill_pair_group(ws3, 6, 14, edu)
    _fill_pair_group(ws3, 6, 22, book)

    _write_final_review_sheet(wb[REVIEW_SHEET], records, meta)
    out = BytesIO(); wb.save(out)
    return out.getvalue(), records, meta, (base, edu, book)
