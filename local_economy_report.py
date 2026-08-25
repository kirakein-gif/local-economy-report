import streamlit as st
import pandas as pd
import openpyxl
import requests
import urllib.parse
import base64
import time
import uuid
from io import BytesIO

# ---------------------------------------------------------
# 1. 환경 설정 및 상수
# ---------------------------------------------------------
st.set_page_config(page_title="지역경제활성화 자동 집계 시스템", page_icon="📊", layout="wide", initial_sidebar_state="expanded")

# ===========================================================
# 0. 동시접속 제한 + 대기열
# -----------------------------------------------------------
# 무료 티어(Streamlit Community Cloud)의 공유 메모리 한도를 넘지 않도록
# 동시 "활성 사용자"를 MAX_CONCURRENT 명으로 제한한다.
# - st.cache_resource로 만든 객체는 이 앱을 쓰는 '모든' 세션이 공유한다
#   (단, 서버 인스턴스가 1대일 때만 유효 — 여러 대로 복제하면 인스턴스마다
#   따로 생기므로 실제 한도가 인스턴스 수만큼 늘어난다는 점은 알아두어야 함)
# - 자리가 다 찼으면 대기열에 줄을 세우고, 화면이 몇 초마다 자동 새로고침
#   되면서 "내 차례가 됐는지" 스스로 확인한다(서버가 밀어주는 방식이 아니라
#   각자 화면이 주기적으로 다시 확인하는 방식 — Streamlit 자체엔 서버 푸시
#   기능이 없어서 이렇게 구현하는 것이 무료 티어 안에서 가장 단순한 방법)
# - IDLE_TIMEOUT 동안 아무 조작(재실행 트리거)이 없으면 자동으로 자리 반납
# - "나가기" 버튼으로 즉시 자리 반납 가능
# ===========================================================
MAX_CONCURRENT = 10
IDLE_TIMEOUT_SECONDS = 600      # 10분간 조작 없으면 자동 퇴장
QUEUE_POLL_SECONDS = 5          # 대기 화면 자동 새로고침 주기


@st.cache_resource
def get_room_state():
    # 모든 세션이 공유하는 객체. active: {세션ID: 마지막 활동시각}
    # waiting: {세션ID: (대기열 진입시각, 마지막으로 확인된 시각)}
    return {"active": {}, "waiting": {}}


WAITING_STALE_SECONDS = QUEUE_POLL_SECONDS * 6  # 이 시간 이상 새로고침이 없으면 이탈한 것으로 간주


def _acquire_slot() -> bool:
    """이 세션이 앱을 사용할 수 있는 상태인지 확인/갱신하고, 사용 가능하면 True를 반환한다."""
    room = get_room_state()
    now = time.time()
    my_sid = st.session_state.session_id

    # 1) 오래(IDLE_TIMEOUT_SECONDS) 활동이 없는 활성 세션은 자동 퇴장 처리
    expired_active = [sid for sid, t in room["active"].items()
                       if sid != my_sid and now - t > IDLE_TIMEOUT_SECONDS]
    for sid in expired_active:
        del room["active"][sid]

    # 1-1) 새로고침이 한동안 없었던 대기열 항목(탭을 닫고 떠난 경우)도 정리
    stale_waiting = [sid for sid, (_, seen) in room["waiting"].items()
                      if sid != my_sid and now - seen > WAITING_STALE_SECONDS]
    for sid in stale_waiting:
        del room["waiting"][sid]

    # 2) 나는 이미 활성 상태 → 활동시각만 갱신하고 통과
    if my_sid in room["active"]:
        room["active"][my_sid] = now
        room["waiting"].pop(my_sid, None)
        return True

    # 3) 아직 활성이 아니면 대기열에 등록.
    #    이미 등록돼 있으면 '진입 순서(joined)'는 그대로 두고 '마지막 확인 시각'만 갱신 —
    #    그래야 5초마다 자동 새로고침돼도 줄 순서가 뒤로 밀리지 않는다.
    joined = room["waiting"].get(my_sid, (now, now))[0]
    room["waiting"][my_sid] = (joined, now)

    # 4) 대기열 앞쪽(진입이 빠른 순)부터, 남는 자리만큼 활성으로 승격
    waiting_order = sorted(room["waiting"].items(), key=lambda x: x[1][0])
    free_slots = MAX_CONCURRENT - len(room["active"])
    for sid, _ in waiting_order[:max(free_slots, 0)]:
        room["active"][sid] = now
        room["waiting"].pop(sid, None)

    return my_sid in room["active"]


def _render_waiting_room():
    room = get_room_state()
    my_sid = st.session_state.session_id
    waiting_order = sorted(room["waiting"].items(), key=lambda x: x[1][0])
    my_position = next((i for i, (sid, _) in enumerate(waiting_order) if sid == my_sid), None)
    position_txt = f"{my_position + 1}번째" if my_position is not None else "확인 중"

    st.markdown("""
    <style>
        body { font-family: 'Pretendard', -apple-system, sans-serif; }
    </style>
    <meta http-equiv="refresh" content="%d">
    """ % QUEUE_POLL_SECONDS, unsafe_allow_html=True)

    st.markdown(f"""
    <div style="max-width:480px; margin:80px auto; text-align:center; padding:36px;
                border:1px solid #E6E9EF; border-radius:16px; background:#fff;">
        <div style="font-size:2rem; margin-bottom:12px;">⏳</div>
        <div style="font-size:1.1rem; font-weight:700; color:#16203A; margin-bottom:8px;">
            현재 사용자가 많아 대기 중입니다
        </div>
        <div style="color:#6B7280; font-size:0.9rem; margin-bottom:18px;">
            대기 순번: <b>{position_txt}</b> · 자리가 나면 자동으로 접속됩니다.<br>
            (화면이 {QUEUE_POLL_SECONDS}초마다 자동으로 새로고침됩니다 — 그냥 기다리시면 됩니다)
        </div>
        <div style="font-size:0.78rem; color:#93A2C4;">
            최대 동시 사용자 {MAX_CONCURRENT}명 · 현재 사용 중 {len(room['active'])}명
        </div>
    </div>
    """, unsafe_allow_html=True)
    st.stop()


# 세션 고유 ID — URL 쿼리 파라미터(?sid=...)에 저장해서 자동 새로고침으로 페이지가
# 완전히 다시 로드되어도(= 새 웹소켓 연결 = 새 session_state) 같은 ID를 계속 쓰게 한다.
# session_state에만 저장하면 새로고침마다 ID가 바뀌어 대기열 순서가 계속 초기화되는
# 문제가 있었다.
if "sid" in st.query_params:
    st.session_state.session_id = st.query_params["sid"]
else:
    st.session_state.session_id = str(uuid.uuid4())
    st.query_params["sid"] = st.session_state.session_id

if not _acquire_slot():

    _render_waiting_room()

# ===========================================================
# 디자인 시스템
# -----------------------------------------------------------
# 톤앤매너: 국내 핀테크(토스/뱅크샐러드류) 스케일업 기업의
#          내부 정산·리포팅 대시보드를 참고한 신뢰감 있는 톤.
# 컬러 토큰
#   --ink       #0B1220  브랜드/사이드바 배경 (딥 네이비)
#   --ink-2     #131B2E  사이드바 카드 배경
#   --blue      #2F6FED  주 액션(프라이머리) 블루
#   --mint      #00C2A8  성공/완료 상태
#   --amber     #F59E0B  경고/미완료 상태
#   --bg        #F5F7FA  본문 배경
#   --card      #FFFFFF  카드 배경
#   --line      #E6E9EF  구분선
#   --text      #16203A  본문 텍스트
#   --text-2    #6B7280  보조 텍스트
# 타이포그래피: Pretendard(제목/본문) + JetBrains Mono(수치/금액)
# 시그니처 요소: 지역별(천안/충남/기타) 실적 비중을 보여주는
#               하단 분포 바 — 이 앱의 핵심 업무(관내구매 집계)를
#               그대로 시각화한 요소.
# ===========================================================
st.markdown("""
<style>
    @import url('https://cdn.jsdelivr.net/gh/orioncactus/pretendard@v1.3.9/dist/web/static/pretendard.css');
    @import url('https://cdn.jsdelivr.net/npm/@fontsource/jetbrains-mono@5.0.3/index.css');

    :root {
        --ink: #0B1220;
        --ink-2: #131B2E;
        --blue: #2F6FED;
        --blue-dark: #1E4FC2;
        --mint: #00C2A8;
        --amber: #F59E0B;
        --bg: #F5F7FA;
        --card: #FFFFFF;
        --line: #E6E9EF;
        --text: #16203A;
        --text-2: #6B7280;
    }

    html, body, [class*="css"] {
        font-family: 'Pretendard', -apple-system, sans-serif;
    }

    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}

    .stApp {
        background-color: var(--bg);
    }

    /* ---------- 사이드바 ---------- */
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, var(--ink) 0%, var(--ink-2) 100%) !important;
    }
    [data-testid="stSidebar"] p,
    [data-testid="stSidebar"] h1,
    [data-testid="stSidebar"] h2,
    [data-testid="stSidebar"] h3,
    [data-testid="stSidebar"] label,
    [data-testid="stSidebar"] .stMarkdown {
        color: #ffffff !important;
        letter-spacing: -0.01em;
    }
    [data-testid="stSidebar"] .stSelectbox label,
    [data-testid="stSidebar"] .stFileUploader label {
        font-size: 0.82rem;
        font-weight: 600;
        color: #93A2C4 !important;
        text-transform: uppercase;
        letter-spacing: 0.04em;
    }
    [data-testid="stSidebar"] h2 {
        font-size: 0.95rem;
        font-weight: 700;
        border-bottom: 1px solid rgba(255,255,255,0.12);
        padding-bottom: 10px;
        margin-top: 4px;
    }

    [data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"],
    [data-testid="stSidebar"] [data-testid="stFileUploadDropzone"] {
        background-color: #ffffff !important;
        border-radius: 10px !important;
        border: 1px dashed #C7D2E8 !important;
    }
    [data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] *,
    [data-testid="stSidebar"] [data-testid="stFileUploadDropzone"] * {
        color: #333333 !important;
        fill: #333333 !important;
    }
    /* 업로드 박스 안의 '찾아보기(Browse files)' 버튼 — 배경색과 겹쳐 글자가
       안 보이던 부분을 별도 배경/테두리로 분리해서 확실히 보이게 함 */
    [data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button,
    [data-testid="stSidebar"] [data-testid="stFileUploadDropzone"] button,
    [data-testid="stSidebar"] [data-testid="stBaseButton-secondary"] {
        background-color: #EEF1F6 !important;
        border: 1px solid #D8DEE9 !important;
        border-radius: 8px !important;
    }
    [data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button *,
    [data-testid="stSidebar"] [data-testid="stFileUploadDropzone"] button *,
    [data-testid="stSidebar"] [data-testid="stBaseButton-secondary"] * {
        color: #16203A !important;
        fill: #16203A !important;
    }
    [data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] small,
    [data-testid="stSidebar"] [data-testid="stFileUploadDropzone"] small {
        color: #6B7280 !important;
    }
    /* 업로드된 개별 파일의 삭제 아이콘 — 흐린 회색 원형이라 잘 안 보이던 것을
       뚜렷한 빨간색 X로 강조 (Streamlit 버전에 따라 testid/aria-label이 달라
       여러 선택자를 함께 걸어 둠) */
    [data-testid="stSidebar"] [data-testid="stFileUploaderDeleteBtn"],
    [data-testid="stSidebar"] [data-testid="stFileUploaderFile"] button,
    [data-testid="stSidebar"] button[aria-label*="Remove" i],
    [data-testid="stSidebar"] button[title*="Remove" i] {
        opacity: 1 !important;
    }
    [data-testid="stSidebar"] [data-testid="stFileUploaderDeleteBtn"] svg,
    [data-testid="stSidebar"] [data-testid="stFileUploaderFile"] button svg,
    [data-testid="stSidebar"] button[aria-label*="Remove" i] svg,
    [data-testid="stSidebar"] button[title*="Remove" i] svg {
        fill: #E4572E !important;
        stroke: #E4572E !important;
        opacity: 1 !important;
    }
    .upload-note {
        font-size: 0.76rem;
        color: #93A2C4;
        margin-top: 6px;
        line-height: 1.4;
    }
    .amount-input-row {
        display: flex;
        align-items: center;
        gap: 8px;
    }

    [data-testid="collapsedControl"],
    [data-testid="stSidebarCollapseButton"],
    button[title="Collapse sidebar"],
    button[title="View sidebar"] {
        display: none !important;
    }

    /* ---------- 브랜드 헤더 ---------- */
    .brand-header {
        display: flex;
        align-items: center;
        gap: 14px;
        padding-bottom: 20px;
        margin-bottom: 24px;
        border-bottom: 1px solid var(--line);
    }
    .brand-mark {
        width: 42px; height: 42px;
        border-radius: 11px;
        background: linear-gradient(135deg, var(--blue), #5B8DF6);
        display: flex; align-items: center; justify-content: center;
        font-size: 20px;
        box-shadow: 0 6px 16px rgba(47,111,237,0.28);
        flex-shrink: 0;
    }
    .brand-title {
        font-size: 1.5rem;
        font-weight: 800;
        color: var(--text);
        letter-spacing: -0.02em;
        line-height: 1.2;
    }
    .brand-sub {
        font-size: 0.85rem;
        color: var(--text-2);
        margin-top: 2px;
    }

    /* ---------- KPI 카드 ---------- */
    .kpi-row { display: flex; gap: 14px; margin-bottom: 22px; }
    .kpi-card {
        flex: 1;
        background: var(--card);
        border: 1px solid var(--line);
        border-radius: 14px;
        padding: 18px 20px;
        box-shadow: 0 1px 2px rgba(15,23,42,0.03);
    }
    .kpi-label {
        font-size: 0.78rem;
        color: var(--text-2);
        font-weight: 600;
        margin-bottom: 8px;
    }
    .kpi-value {
        font-family: 'JetBrains Mono', monospace;
        font-size: 1.6rem;
        font-weight: 700;
        color: var(--text);
        letter-spacing: -0.02em;
    }
    .kpi-delta {
        font-size: 0.78rem;
        font-weight: 600;
        margin-top: 6px;
        display: inline-block;
        padding: 2px 8px;
        border-radius: 6px;
    }
    .kpi-delta.ok { color: var(--mint); background: rgba(0,194,168,0.12); }
    .kpi-delta.warn { color: var(--amber); background: rgba(245,158,11,0.12); }

    /* ---------- 지역 분포 바 (시그니처 요소) ---------- */
    .region-bar-card {
        background: var(--card);
        border: 1px solid var(--line);
        border-radius: 14px;
        padding: 18px 20px;
        margin-bottom: 22px;
    }
    .region-bar-title {
        font-size: 0.82rem;
        font-weight: 700;
        color: var(--text);
        margin-bottom: 12px;
    }
    .region-bar-track {
        width: 100%; height: 10px;
        border-radius: 6px;
        overflow: hidden;
        display: flex;
        background: #EEF1F6;
    }
    .region-bar-seg { height: 100%; }
    .region-legend {
        display: flex; gap: 18px; margin-top: 10px;
        font-size: 0.78rem; color: var(--text-2);
    }
    .region-legend span.dot {
        display: inline-block; width: 8px; height: 8px;
        border-radius: 50%; margin-right: 6px;
    }

    /* ---------- 버튼 ---------- */
    div.stButton > button {
        background-color: var(--blue);
        color: white;
        border-radius: 10px;
        border: none;
        padding: 10px 22px;
        font-weight: 600;
        font-size: 0.9rem;
        transition: all 0.2s ease;
        box-shadow: 0 1px 2px rgba(15,23,42,0.06);
    }
    div.stButton > button:hover {
        background-color: var(--blue-dark);
        transform: translateY(-1px);
        box-shadow: 0 6px 16px rgba(47,111,237,0.28);
    }

    div.stDownloadButton > button {
        background-color: var(--ink);
        color: white;
        border-radius: 10px;
        border: none;
        font-weight: 600;
        font-size: 0.9rem;
        transition: all 0.2s ease;
        width: 100%;
    }
    div.stDownloadButton > button:hover {
        background-color: var(--blue-dark);
        transform: translateY(-1px);
    }

    /* ---------- 상태 배지 ---------- */
    .status-pill {
        display: inline-flex;
        align-items: center;
        gap: 6px;
        padding: 4px 12px;
        border-radius: 999px;
        font-size: 0.8rem;
        font-weight: 600;
    }
    .status-pill.warn { background: rgba(245,158,11,0.12); color: #B45309; }
    .status-pill.ok { background: rgba(0,194,168,0.12); color: #047857; }

    .main-title { display: none; } /* 구 타이틀 스타일 비활성화 */
</style>
""", unsafe_allow_html=True)

MY_G2B_API_KEY = "V%2FBFQCvaQlP%2F3ebvSQyuncyYbTzwqIxQ5yDO%2Fc%2FnX3YTRLd3ZZXxTeNhVd99xGMLoQOWLSwS7x%2BJ07aIn7Fk0w%3D%3D"
API_URL = "https://apis.data.go.kr/1230000/ao/UsrInfoService02/getPrcrmntCorpBasicInfo02"
CHUNGNAM_REGIONS = ["천안", "아산", "공주", "보령", "서산", "논산", "계룡", "당진", "금산", "부여", "서천", "청양", "홍성", "예산", "태안"]
DEFAULT_TARGET_AMOUNT = 500000  # 사이드바에서 값을 바꾸지 않았을 때 쓰는 기본 금액기준

# ★ 변환하신 템플릿의 긴 문자열을 아래 따옴표 사이에 반드시 넣어주십시오.
TEMPLATE_BASE64 = "UEsDBBQABgAIAAAAIQB0NlqmegEAAIQFAAATAAgCW0NvbnRlbnRfVHlwZXNdLnhtbCCiBAIooAACAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACsVM1OAjEQvpv4DpteDVvwYIxh4YB6VBLwAWo7sA3dtukMCG/vbEFiDEIIXLbZtvP9TGemP1w3rlhBQht8JXplVxTgdTDWzyvxMX3tPIoCSXmjXPBQiQ2gGA5ub/rTTQQsONpjJWqi+CQl6hoahWWI4PlkFlKjiH/TXEalF2oO8r7bfZA6eAJPHWoxxKD/DDO1dFS8rHl7q+TTelGMtvdaqkqoGJ3VilioXHnzh6QTZjOrwQS9bBi6xJhAGawBqHFlTJYZ0wSI2BgKeZAzgcPzSHeuSo7MwrC2Ee/Y+j8M7cn/rnZx7/wcyRooxirRm2rYu1w7+RXS4jOERXkc5NzU5BSVjbL+R/cR/nwZZV56VxbS+svAJ3QQ1xjI/L1cQoY5QYi0cYDXTnsGPcVcqwRmQly986sL+I19QodWTo9qLpErJ2GPe4yfW3qcQkSeGgnOF/DTom10JzIQJLKwb9JDxb5n5JFzsWNoZ5oBc4Bb5hk6+AYAAP//AwBQSwMEFAAGAAgAAAAhALVVMCP0AAAATAIAAAsACAJfcmVscy8ucmVscyCiBAIooAACAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACskk1PwzAMhu9I/IfI99XdkBBCS3dBSLshVH6ASdwPtY2jJBvdvyccEFQagwNHf71+/Mrb3TyN6sgh9uI0rIsSFDsjtnethpf6cXUHKiZylkZxrOHEEXbV9dX2mUdKeSh2vY8qq7iooUvJ3yNG0/FEsRDPLlcaCROlHIYWPZmBWsZNWd5i+K4B1UJT7a2GsLc3oOqTz5t/15am6Q0/iDlM7NKZFchzYmfZrnzIbCH1+RpVU2g5abBinnI6InlfZGzA80SbvxP9fC1OnMhSIjQS+DLPR8cloPV/WrQ08cudecQ3CcOryPDJgosfqN4BAAD//wMAUEsDBBQABgAIAAAAIQDHeXvKwAMAADAJAAAPAAAAeGwvd29ya2Jvb2sueG1srFXbbuM2EH0v0H9QibwqEnWzJMReSLaEBogXgeMmW2ABg5HoiIhupajYQbBv+9ZfSIF+Rb8qH9GhbDnOuijcbA2bNMnR4ZmZM6OzD+siVx4ob1hVDhE+1ZFCy6RKWXk3RL/MY9VFSiNImZK8KukQPdIGfRj9+MPZquL3t1V1rwBA2QxRJkTta1qTZLQgzWlV0xJOlhUviIAlv9OamlOSNhmlosg1Q9cdrSCsRBsEnx+DUS2XLKGTKmkLWooNCKc5EUC/yVjd9GhFcgxcQfh9W6tJVdQAcctyJh47UKQUiX9+V1ac3Obg9hrbyprD14Ef1mEw+pvg6OCqgiW8aqqlOAVobUP6wH+saxi/CcH6MAbHIVkapw9M5nDHijvvZOXssJxXMKx/NxoGaXVa8SF470Szd9wMNDpbspxeb6SrkLr+SAqZqRwpOWlElDJB0yEawLJa0TcbvK3DluVwauqG4SJttJPzJVdSuiRtLuYg5B4eKsNxPMOWliCMIBeUl0TQcVUK0OHWr+/VXIc9zipQuDKjv7WMUygs0Bf4CiNJfHLbXBKRKS3Ph+iT/7l+/LwnQ3Ko+f8gRJJI7zRwb0Nh8/9bV4EJ93uxXQquwP/zyQUE/Io8QPghyem2Os8hvthclAn38eIp9CJ7bEaWGgUBVi3Li9RQDyI10D0rdkPD8nD0BZzhjp9UpBXZNrMSeogsSOPB0ZSs+xOs+y1LX2k86duPKudvhv7si3RY9rBrRlfNqwbkUlnfsDKtVkOkYsd0bKQ89hsDF5xcdac3LBUZeGnoA7DY7P1M2V0GlLFtGVLy3JDUhujJjkzLcz1TdT1PVy0jNNRwYA7UcaDrVjgeh7FrdJS0PU5duwRu3ayUncRf/nx++ev55evzy+9/QHeWDbWLNVK4L2/i5ynuctk/DIpmJU1lgQDU3moLuFjnZXF6yVkpFgE0aVkyCcmvemQdjfbv/OkkOMH+yfQE4zNtDw0U8/YmwEigoOTUEfSwbniSGV2Li0Z0M2iZQXCwpQcDkIKqR6atWq5nqK5lGurYmhiRPYgmUWhLeciXjf9/tNyupPz+LSZZZoSLOSfJPbz7ZnQZkgb0vAkk8N0nG9puqJtA0YpxrFoY0hmGjqXak9i0B3gyjuz4lax0f/nOhudq3dOUiBaagewD3dqXY7zd3W0uNxvbpL4pfX82kXHfPv1vhlfgfU6PNI6vjzQcf5zOp0faXkTzxU18rHEwDSfB8fbBbBb8Oo8+9Vdo/xhQrUu4HDuZar1MRn8DAAD//wMAUEsDBBQABgAIAAAAIQCSB5TsBAEAAD8DAAAaAAgBeGwvX3JlbHMvd29ya2Jvb2sueG1sLnJlbHMgogQBKKAAAQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACskstqxDAMRfeF/oPRvnEyfVCGcWbRUphtm36AcJQ4TGIHW33k72tSOsnAkG6yMUjC9x6Ju9t/d634JB8aZxVkSQqCrHZlY2sF78XLzSOIwGhLbJ0lBQMF2OfXV7tXapHjp2CaPoioYoMCw9xvpQzaUIchcT3ZOKmc75Bj6WvZoz5iTXKTpg/SzzUgP9MUh1KBP5S3IIqhj87/a7uqajQ9O/3RkeULFjLw0MYFRIG+JlbwWyeREeRl+82a9hzPQpP7WMrxzZYYsjUZvpw/BkPEE8epFeQ4WYS5XxNGY6ufDDZ2gjm1li5yt2ooDHoq39jHzM+zMW//wciz2Oc/AAAA//8DAFBLAwQUAAYACAAAACEAmVifP/EEAACWDwAAGAAAAHhsL3dvcmtzaGVldHMvc2hlZXQxLnhtbJxXW4/qNhB+r9T/EOWpfYCQBAJEwNFCyAILVVXOaZ9NMBCRxKljdpdz1P/emSQEsNkDWmkv5mPmG8/F43Hvy3scaa+UZyFL+rpZb+gaTQK2DpNtX//21a91dC0TJFmTiCW0rx9ppn8Z/PpL743xfbajVGjAkGR9fSdE6hpGFuxoTLI6S2kC32wYj4mAj3xrZCmnZJ0rxZFhNRqOEZMw0QsGlz/CwTabMKAeCw4xTURBwmlEBOw/24VpdmKLg0foYsL3h7QWsDgFilUYheKYk+paHLjTbcI4WUXg97vZJIH2zuHHgl/7ZCbHFUtxGHCWsY2oA7NR7Fl1v2t0DRJUTKr/D9GYTYPT1xATeKayPrcls1VxWWcy+5NkTkWG4eLuIVz39R+2P/KazrBT87ueU2uafqfWdWyvNuxYo67XbjTG1vA/fdBbh5Bh9ErjdNPXn0x3YZq6MejlBfR3SN+yi7UmyGpJIxoICkZMXfvOWLwMSET/wAqMAGtAbVfoEkt3To7sIJCq/BqLesXYHqEp8DRgH1nOivsggQhf6YhGwDY2gS37N98armFfRrWxy/Vpk35+EP7k2ppuyCESf7G3CQ23OwGmW/V2CyKEJeaujx7NAqhtsF63kTdgEZDAXy0O8YxCaZJ3+A9BfQvXYgdLq+5YwJCJI9YqSASHTLD4n/LrkqXUb54IHLsD65Kje6GOzhRWc5c8Isigx9mbBiUK5FlK8MCbLm4BHbCcOpq/6QDsHNWeUA93qmugkEFcXwfdnvGKlkqRYSWCTqPSSEE8BRkriK8gzwoyUZCpgswU5EVB5gqyKBArrwiIWRU48FwJnOnUIW5iFwb7IcNKuB3EFpKV8YeQqzTt+/FHPSy1vBYwuMMSgX9VRhrXGRlVIqeMeAoyvkHTuabxb4hY1yLPCu9EQab3Lc1uiJjXll4U3rmCLEoEjngVmrNPxmVW4QCp6YA784Gcnk4GUkBmnHNmCgROZGXdlhKjSjSvJTxVwpSyOy4NNyvD/l3Dz3cNT+4bniqGZ3cNv9w1PL9veFEazq+RqzRi85S72sXh/MrSDw8ntLKiwyFHX3cusuZIDa6QaFfxHsmAJwPjEzDobQbLb4vfhi3Xa/3eMzbYQKWE+rL2swxMZGAq8/std/IR/0zWfpGBuQwsCqADQSr2P2u58xv8V8mAEH6+U56ygSR9/bK5taVsFBLn6h/JgCcD4xNQZcNxPeejbMjazzIwkYGpzO877uQj/pms/SIDcxlYFEDrnA3Hnd/gv8oGPgAuj8ZPb3kUzk9K1bnMc6e/osV542FaFJZoz239ijYfyx7mzaUl4nOrvSaWB5+fTzvlXHDRDFpV+RW8xbhYzFYx5dt8sMy0gB1w9sMTU6HFBLywXWhgcBNLOEzGMIao+Nh2ocWr+NR2oQPf4LHdp1v40Hbh2lflfduFS1rFZ7YLVyrOkGe3Br10B+9HEQYwAW9YInC6hhiJYwoDa8JGLCkfoaiYki1dEL4Nk0yL6CYfhdu6xotxuVGHtcCODCvoLysmYNY9fdrBE5PCeIPDM1hi4vSh5F1ScUi1lKSUL8PvYBxrC98Jfb0JS8ZDmL3z52RfTxkXnIQCTLv4duHTdR5omI2H8JTdV9kCT2KSHEgBj4oUmpDCFd9rqImTcj61W3AFgSg6X07ZBRNEq3pRD/4HAAD//wMAUEsDBBQABgAIAAAAIQB1ro+8aQcAAAMhAAATAAAAeGwvdGhlbWUvdGhlbWUxLnhtbOxZ3YscNxJ/P7j/QfT7eL6652PxOMynN/aubbxjhzxqZzTT8qpbg6TZ9RAMweHgAuHgIAl5CeQtDyFc4AIX7uX+GIPNne/+hyupe6alHU127ayPXNhd2J3W/KpUqir9VF26+d7ThKFTIiTlaSeo3qgEiKQTPqXpvBM8Go9KrQBJhdMpZjwlnWBFZPDerd//7ibeUzFJCAL5VO7hThArtdgrl+UEhrG8wRckhe9mXCRYwaOYl6cCn4HehJVrlUqjnGCaBijFCai9P5vRCUG1SrWOSvCvVkP/+cNnr7/9Y3BrPdOQwXSpknpgwsSRnofsFDdy05OqRsuV7DOBTjHrBGDAlJ+NyVMVIIalgi86QcX8BOVbN8t4LxdiaoesJTcyP7lcLjA9qZk5xfx4M2kYRmGju9FvAExt44bNYWPY2OgzADyZwKozW1ydzVo/zLEWKPvo0T1oDupVB2/pr2/Z3I30r4M3oEx/uIUfjfrgRQdvQBk+2sJHvXZv4Oo3oAzf2MI3K91B2HT0G1DMaHqyha5EjXp/vdoNZMbZvhfejsJRs5YrL1CQDZtM01PMeKouk3cJfsLFCMBaiGFFU6RWCzLDE0j1Pmb0WFB0QOcxJOECp1zCcKVWGVXq8Ff/huaTiS7eI9iS1jaCVXJrSNuG5ETQheoEd0BrYEFe/vTTi+c/vnj+txeffPLi+V/yuY0qR24fp3Nb7vW3f/731x+jf/31m9eff5FNfR4vbfyr7z999fd//Jx6WHHhipdf/vDqxx9efvWnf373uUd7V+BjGz6mCZHoHjlDD3kCC/TYT47Fm0mMY0wdCRyDbo/qoYod4L0VZj5cj7gufCyAcXzA28snjq1HsVgq6pn5bpw4wEPOWY8LrwPu6rksD4+X6dw/uVjauIcYn/rm7uPUCfBwuQDapT6V/Zg4Zj5gOFV4TlKikP6OnxDiWd2HlDp+PaQTwSWfKfQhRT1MvS4Z02MnkQqhfZpAXFY+AyHUjm8OH6MeZ75VD8ipi4RtgZnH+DFhjhtv46XCiU/lGCfMdvgBVrHPyKOVmNi4oVQQ6TlhHA2nREqfzH0B67WCfhcYxh/2Q7ZKXKRQ9MSn8wBzbiMH/KQf42ThtZmmsY19X55AimL0gCsf/JC7O0Q/QxxwujPcjylxwn0xETwCcrVNKhJEf7MUnljeJtzdjys2w8THMl2ROOzaFdSbHb3l3EntA0IYPsNTQtCj9z0W9PjC8Xlh9J0YWGWf+BLrDnZzVT+nRBJkapxtijyg0knZIzLnO+w5XJ0jnhVOEyx2ab4HUXdSF045L5XeZ5MTG3iPQo0I+eJ1yn0JOqzkHu7S+iDGztmln6U/X1fCid9l9hjsyydvui9BhryxDBD7pX0zxsyZoEiYMYYCw0e3IOKEvxDR56oRW3rlZu6mLcIARZJT7yQ0vbD4OVf2RP+bssdfwFxBweNX/EtKnV2Usn+uwNmF+z8sawZ4mT4gcJJsc9Z1VXNd1QS/+apm116+rmWua5nrWsb39vVOapmifIHKpuj4mP5Pcqn2z4wydqRWjBxI0wGS8HYzHcGgaVOZvuWmNbiI4WPeeHJwc4GNDBJcfUBVfBTjBbSJqqaxOZe56rlECy6he2SGTe+VnNNtelDL5JBPsw5otaq7nZk7JVbFeCXajEPHSmXoRrPo6m3Umz7p3HRi1wZo2TcxwprMNaLuMaK5HoSI/JwRZmVXYkXbY0VLq1+Hah3FjSvAtE1U4PUbwUt7J4jCrLMMjTko1ac6TlmTeR1dHZwrjfQuZzI7A6DcXmdAEem2tnXn8vTqslS7RKQdI6x0c42w0jCGl+I8O+1W/FXGul2E1DFPu2K9Gwozmq13EWtNKOe4gaU2U7AUnXWCRj2Ca5gJXnSCGXSP4WOygNyR+g0Msznc00yUyDb82zDLQkg1wDLOHG5IJ2ODhCoiEKNJJ9DL32QDSw2HGNuqNSCEX61xbaCVX5txEHQ3yGQ2IxNlh90a0Z7OHoHhM67wfmvE3x6sJfkSwn0UT8/QMVuKhxhSLGpWtQOnVMIlQjXz5pTCDdmGyIr8O3cw5bRrX1GZHMrGMVvEOD9RbDLP4IZEN+aYp40PrKd8zeDQbRcez/UB+4tP3YuPau05izSLM9NhFX1q+sn03R3yllXFIepYlVG3eb+WBde111wHieo9JS44dS9xIFimFZM5pmmLt2lYc3Y+6pp2hQWB5YnGDr9tzgivJ9725Ae581mrD4h1jWkS39yx2zff/PgJkMcA7hKXTEkTSrjLFhiKvuxmMqMN2CJPVV4jwie0FLQTfFSJumG/FvVLlVY0LIX1sFJqRd16qRtF9eowqlYGvdozOFhUnFSj7H5/BNcZbJXf8pvxrZv+ZH1jc2PCkzI3N/llY7i56a/WLrrpH+ub/ABRIKCPGrVRu97uNUrtendUCge9Vqndb/RKg0a/ORgN+lGrPXoWoFMDDrv1ftgYtkqNar9fChsVvZRWu9QMa7Vu2Oy2hmH3WV7SgBcyKsn9Aq42Nt76LwAAAP//AwBQSwMEFAAGAAgAAAAhALyV9S/KBAAALB8AAA0AAAB4bC9zdHlsZXMueG1s5FnNbuM2EL4X6DsIujv6seXYhuXFOomABbZFgaRAr7RE2cRSpEHRWXuLAjku0EOxQPfUFmiBAu2ht/apmuw7dEhJlpxEjvO79jY5WKLIme+bGQ6HZP/ZPKHGKRYp4cw3nT3bNDALeUTY2De/PgkaHdNIJWIRopxh31zg1Hw2+PyzfioXFB9PMJYGiGCpb06knPYsKw0nOEHpHp9iBl9iLhIk4VWMrXQqMIpSNSihlmvbbStBhJmZhF4SbiIkQeLVbNoIeTJFkowIJXKhZZlGEvZejBkXaEQB6txpodCYO23hGnNRKNGtV/QkJBQ85bHcA7kWj2MS4qtwu1bXQmEpCSTfTZLjWba7wn0u7iipZQl8SpT7zEE/5kymRshnTPrmPgBVJui9Yvw1C9Qn8HDea9BP3xiniEKLY1qDfsgpF4YE14HldAtDCc56nP/x7uKXM+Pfv389//G96hyjhNBF9tHVoydIpBAJmUC3q9p0HOQSEgJeUY2WQpjhLBF01JenU3cz4Q/v/7z45+35b3+d//D9xU9nlyk36yivsBspGxQ2zqykbSzGI98MAlv/rTJ/eL0V3z623uvi6Um56jBaF8f3NK/2bgrBSyhdTrKmmk/QMOhDNpJYsABejPz5ZDGF2cQgcWaxr/vd0Hss0MJxvc0HpJySSKEYH1Tj68hW/0rMKP9AWITnOPLNdktLrwBWE1OD0z/AccRFBItCkUocRTNrG/QpjiWIFWQ8Ub+ST5USLiVkzkE/ImjMGaJqshcjqiNhNYGFwzcTHJFZAmIzj10Gp5TkOooRcgJLRV1/jUaD2VABwC5Qb6QgI7g5v7Vo/3fsbvD21nnv0aNzbXhcE5s3ANqy6NwxdreOzo/M79Gic5nIdyklbhfojWPpcWFvXQKL+Ax2ZbXr944nsI/M7tET2D345VUgFJUhpvRYVX/fxGVlCRExjw02S4JEvoDiFA4g1B6xeISqNH/MisjsZdBHlIxZghnsObGQJFQ72RBecbbNnMdQflb1ZdqrivfvpNmYx7eAUMfMBZLXM1vKN9B0Shdq155t2jcgXKetfVttme7nhYkvA5hwQd4AsIrN671QB6pVAwraCxOvmiB7G+qNiUL0ECCN1wJNT/Bc21jtWFTcrCBuluHo1SCG9lXEX86SERaBPvgqkRauvAcPXSabV239IDQgSraeRiU51MWPOvXKp+iOxA+csH4K8dPdrfhREztPwZCNy8UFTgfXBpA63ykn9VOno8oEqCMA7esmwNbjd8AdO0FgK5fVuqjYFaN+stNy19OKA6lnq6alLu+hoK/sK1Z2Fcvq31BXS7754d3PF7+fVUiMZoRKwrJyXhV/yxEgM5qXexR9ii7VpaLevSy1gEUiHKMZlSfLj75ZPn+hj7hh5uW9viKnXGoRvlk+v1Tn6E5bHclDIfoyhWNv+DVmgvjmt0fD/e7hUeA2Ovaw02g1sdfoesPDhtc6GB4eBl3btQ++q1xt3uNiU9/EQvXrtHophetPkZPNwR+Xbb5Zecng6wsFgF3F3nXb9nPPsRtB03YarTbqNDrtptcIPMc9bLeGR17gVbB7d7wAtS3Hya5SFXivJ0mCKWGFrwoPVVvBSfC6hoRVeMIqr7kH/wEAAP//AwBQSwMEFAAGAAgAAAAhAIOTMhIfAgAAswUAABQAAAB4bC9zaGFyZWRTdHJpbmdzLnhtbKxU32vTUBR+F/wfDvdJhS1d0SElzR4Ggm8+KPga2mwNtDe1uRX3lm2ZdKxCEWO7mUiLw1qo0B9p3UP3D+We/A+eNhnoq/EhIeeE853v3PN9V917V6vCW6NhmxYvsp3tHAODl6yyyQ+L7NXLZ1tPGdhC52W9anGjyI4Mm+1p9++pti2AarldZBUh6gVFsUsVo6bb21bd4PTnwGrUdEFh41Cx6w1DL9sVwxC1qpLP5XaVmm5yBiWryUWR5fMMmtx80zT2k8TOY6aptqmpQovmCzwZq4rQ1HqFOAiz9KIBBxYXz8vEmIE4qhMxbu1bPB2EKZqqrKsTBDlexR9bWRDwaoTdn5kQZj562ThsEAAvbwCXI3nyIwudR4BfO/Jbm6A+4dIHOfMKlBrLcRhftaO5u0UPerdR6Mjv4zRIStLUn0EWJmmfaYit3n/AuWmhN8iIk2nPQ4eUgt0znIV46WDQg2RCwItr7B8D9jvRzAF0fbwIMhH9RUJYi2HTMZsYopsJXjsFkEuXPuXcTUljsKI9P5BtJ6KDHdL7nFLw+mG2doSHp8cFyOfyu/LM/avt2uxK4jclMS48yclhG790AIOQ6u4ONJqGmacmMVM/7HvYXcP3Yi9MV4OBA3LuRLNb2ekDiZPs0UJ/JUcTiBYd9Afx53MqlB8mRCt2J/H7ReyNEnds0E4DdKey498BUu2Agr5Phvs34gpdutpvAAAA//8DAFBLAwQUAAYACAAAACEAO20yS8EAAABCAQAAIwAAAHhsL3dvcmtzaGVldHMvX3JlbHMvc2hlZXQxLnhtbC5yZWxzhI/BisIwFEX3A/5DeHuT1oUMQ1M3IrhV5wNi+toG25eQ9xT9e7McZcDl5XDP5Tab+zypG2YOkSzUugKF5GMXaLDwe9otv0GxOOrcFAktPJBh0y6+mgNOTkqJx5BYFQuxhVEk/RjDfsTZsY4JqZA+5tlJiXkwyfmLG9Csqmpt8l8HtC9Ote8s5H1Xgzo9Uln+7I59Hzxuo7/OSPLPhEk5kGA+okg5yEXt8oBiQet39p5rfQ4Epm3My/P2CQAA//8DAFBLAwQUAAYACAAAACEAg/XT2m4CAACYKAAAJwAAAHhsL3ByaW50ZXJTZXR0aW5ncy9wcmludGVyU2V0dGluZ3MxLmJpbuyYv0scQRTHv7N75/njCrlCxCJs0giSkPVHOkEFk4CXBCEilzLcGhAMgUBAu4NYWVmJpLBOlSaFhcFAAsYipLJMGZCY+x/Me/Mj7J07csdJ0OQNvJ3ZNzNv3nzm1+7Uv9V3y1jDC1SxhJd4iggLmMFjlCm9gmdas4h7eIgx3MEolVumMmVUwEHlQvUdH4bD0/VQoUfrEigUUAkCiiuBomfnIZdhgu2GJDc8DSTz1TJnPaFHXAI20T0NTHS/6RvBCLZLO0Mc7wxtl7J0QJE8D2y7HGf5oBnYMs6Nu65Sk88b1oAvv1NKeTIQx6PaDKc5xBSc3QcR8O6WefO42OBCu+W5cq8VjEWvEySZXWIMzMrxMnGAYLpRn1k5paTupEJtil8UrmldUT+NjpOuvxVKsHCbpxQ4welWePj8iSmjRgbmyNBNEra/RbpJkgFbicejjk8qLT57l1Xf/jrWNXK92Dv++CvaP4y+HPn61sj/mJd1S8H4pPQ+4KuQa9/xqzY04q8QaJOAb1F0shP+vUFQ9mTrknE/Q+D+q5Xl5y1y4SOwE7HNTPX9+VaKZgAnJjevT1fWAc3jVSCd+367amk5VmTxCQEhIASEgBC43ATM1ctReJivFd/eZl/Pnt58XzT/2d4XNWVfzC3WvzNLGA/z6h82vAZTXWN99MjoZ9N/8u9XD2b3I7pDlCAEhMB5BHgN/aS1w3fXP3wXzx4DvOTc9jU+aP7nnYmsON9wG3sB4yI/RjK5UwR6wpMCrgdm6snckLlhCfAel5yYPa7ffk9U7ftXoSQEhIAQEAJCQAgIASEgBITAf0PgNwAAAP//AwBQSwMEFAAGAAgAAAAhAI1dBFipAAAA/gAAABAAAAB4bC9jYWxjQ2hhaW4ueG1sXI5BCsIwEEX3gncIs7epgkWkaRdSwYU7PUBIxyaQTEoSRG9vBFukm4F5//H5dftylj0xRONJwLYogSEp3xsaBNxv580BWEySemk9oYA3Rmib9apW0qqTloZYbqAoQKc0HjmPSqOTsfAjUk4ePjiZ8hsGHseAso8aMTnLd2VZcZcLoKkVCwKuFTCTNwCz38snvP/hCVwmbwZLo1sa3Z/B5+XNBwAA//8DAFBLAwQUAAYACAAAACEAM/HBTlkBAABeAgAAEQAIAWRvY1Byb3BzL2NvcmUueG1sIKIEASigAAEAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAfJLdSsMwGIbPBe+h5LxN+rOhoe1AZUcOBDsUz0LybSu2aUmi3c5EPPQCdo/DezDtttqheJi8b548+Ug8WZeF8wpK55VMkO8R5IDklcjlMkHzbOpeIEcbJgUrKgkJ2oBGk/T8LOY15ZWCO1XVoEwO2rEkqSmvE7QypqYYa76CkmnPNqQNF5UqmbFLtcQ1489sCTggZIxLMEwww3ALdOueiA5IwXtk/aKKDiA4hgJKkEZj3/PxT9eAKvWfB7pk0Cxzs6ntmw66Q7bg+7Bvr3XeF5um8Zqw07D+Pn6c3d53T3Vz2c6KA0pjwSlXwEyl0rkGFePBRju8gmkzs3Ne5CCuNunufbv7fPvafsT4d2hhnfueCMKxNnTvfkwewuubbIrSgAQjl0QuCTIS0SikI/+pvfvkfGu33ygPBv8Txy65cP3LjAQ0GtMgHBCPgLTzPv0R6TcAAAD//wMAUEsDBBQABgAIAAAAIQCpQUtb2gEAALMDAAAQAAgBZG9jUHJvcHMvYXBwLnhtbCCiBAEooAABAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAKRTTW/UMBC9I/Efgu9dZ0tVoZXjqtqCegCx0m57RcaZ7FokdmRPo11OIO2J9sABTqRST3DhhBDiP236H3CSNk0p4gC3+Xh6fvNmzPaWWRoUYJ0yOiLDQUgC0NLESs8jcjR7svWIBA6FjkVqNERkBY7s8fv32MSaHCwqcIGn0C4iC8R8RKmTC8iEG/i29p3E2EygT+2cmiRREg6MPMlAI90Ow10KSwQdQ7yVd4SkZRwV+K+ksZG1Pnc8W+VeMGf7eZ4qKdBPyZ8paY0zCQaPlxJSRvtN5tVNQZ5YhSseMtpP2VSKFMaemCcidcDoTYEdgqhNmwhlHWcFjgqQaGzg1Gtv2w4JXgoHtZyIFMIqodHLqmFt0sRp7tDy6tPZ5duv1Wl5+e4nox7Slpuwj+7HaocPG4AP/gq8euL8++bzOqi+vKkuPm7el8Hm24eqXP//a7XcdnAv47YlM4UpuOfJRFj8g0PbfYcala0/V4IvyupHWa3L6vS8r7Jzp+oBHkys0vhi34K4M1GzFK/tNzVjk+VCr3yji54q/cod5TNzIBCuF367yKYLYSH2N9IdRFdgh37XNq1Jxguh5xBfY+426vM8bv8gH+4Owoehv7xejdGb38Z/AQAA//8DAFBLAQItABQABgAIAAAAIQB0NlqmegEAAIQFAAATAAAAAAAAAAAAAAAAAAAAAABbQ29udGVudF9UeXBlc10ueG1sUEsBAi0AFAAGAAgAAAAhALVVMCP0AAAATAIAAAsAAAAAAAAAAAAAAAAAswMAAF9yZWxzLy5yZWxzUEsBAi0AFAAGAAgAAAAhAMd5e8rAAwAAMAkAAA8AAAAAAAAAAAAAAAAA2AYAAHhsL3dvcmtib29rLnhtbFBLAQItABQABgAIAAAAIQCSB5TsBAEAAD8DAAAaAAAAAAAAAAAAAAAAAMUKAAB4bC9fcmVscy93b3JrYm9vay54bWwucmVsc1BLAQItABQABgAIAAAAIQCZWJ8/8QQAAJYPAAAYAAAAAAAAAAAAAAAAAAkNAAB4bC93b3Jrc2hlZXRzL3NoZWV0MS54bWxQSwECLQAUAAYACAAAACEAda6PvGkHAAADIQAAEwAAAAAAAAAAAAAAAAAwEgAAeGwvdGhlbWUvdGhlbWUxLnhtbFBLAQItABQABgAIAAAAIQC8lfUvygQAACwfAAANAAAAAAAAAAAAAAAAAMoZAAB4bC9zdHlsZXMueG1sUEsBAi0AFAAGAAgAAAAhAIOTMhIfAgAAswUAABQAAAAAAAAAAAAAAAAAvx4AAHhsL3NoYXJlZFN0cmluZ3MueG1sUEsBAi0AFAAGAAgAAAAhADttMkvBAAAAQgEAACMAAAAAAAAAAAAAAAAAECEAAHhsL3dvcmtzaGVldHMvX3JlbHMvc2hlZXQxLnhtbC5yZWxzUEsBAi0AFAAGAAgAAAAhAIP109puAgAAmCgAACcAAAAAAAAAAAAAAAAAEiIAAHhsL3ByaW50ZXJTZXR0aW5ncy9wcmludGVyU2V0dGluZ3MxLmJpblBLAQItABQABgAIAAAAIQCNXQRYqQAAAP4AAAAQAAAAAAAAAAAAAAAAAMUkAAB4bC9jYWxjQ2hhaW4ueG1sUEsBAi0AFAAGAAgAAAAhADPxwU5ZAQAAXgIAABEAAAAAAAAAAAAAAAAAnCUAAGRvY1Byb3BzL2NvcmUueG1sUEsBAi0AFAAGAAgAAAAhAKlBS1vaAQAAswMAABAAAAAAAAAAAAAAAAAALCgAAGRvY1Byb3BzL2FwcC54bWxQSwUGAAAAAA0ADQBkAwAAPCsAAAAA"

# ---------------------------------------------------------
# 2. 로직: API 주소 가져오기 (수정본 — inqryDiv=3 사업자등록번호 기준검색)
# ---------------------------------------------------------
def get_addr_api(biz_num):
    """조달청 나라장터 사용자정보서비스에서 사업자등록번호로 업체 주소를 조회한다.

    핵심: inqryDiv=3 (사업자등록번호 기준검색) 일 때만 bizno가 검색조건으로
    사용된다. inqryDiv=1/2(등록일/변경일 기준)는 날짜범위가 필수이며
    bizno는 쓰이지 않으므로, 사업자번호로 찾으려면 반드시 inqryDiv=3이어야 한다.
    """
    if pd.isna(biz_num):
        return None
    biz_clean = str(biz_num).replace("-", "").strip()

    params = {
        'serviceKey': urllib.parse.unquote(MY_G2B_API_KEY),
        'pageNo': '1',
        'numOfRows': '10',
        'type': 'json',
        'inqryDiv': '3',   # ★ 사업자등록번호 기준검색
        'bizno': biz_clean,
    }

    try:
        res = requests.get(API_URL, params=params, timeout=10)
        data = res.json()

        # 조달청 자체 에러 응답 형식 (nkoneps.com.response.ResponseError)
        if "nkoneps.com.response.ResponseError" in data:
            return None

        body = data.get('response', {}).get('body', {})
        items = body.get('items')
        if not items:
            return None

        # items가 배열([...])인 경우와 {"item": [...]}인 경우 모두 대응
        item_list = items if isinstance(items, list) else items.get('item', [])
        if isinstance(item_list, dict):
            item_list = [item_list]

        if item_list:
            first = item_list[0]
            return f"{first.get('adrs', '')} {first.get('dtlAdrs', '')}".strip()

    except Exception:
        pass

    return None


# ---------------------------------------------------------
# 2-1. 캐싱 (리소스 절약 — Streamlit 무료 리소스 증설 신청의 전제조건이기도 함)
# ---------------------------------------------------------
@st.cache_data(show_spinner=False)
def load_excel_normalized(file_bytes: bytes, min_cols: int = 21) -> pd.DataFrame:
    """엑셀 파일을 읽어 컬럼을 위치(0,1,2...) 기준으로 통일한다.
    같은 파일 내용이면 재계산 없이 캐시된 결과를 그대로 재사용한다."""
    d = pd.read_excel(BytesIO(file_bytes))
    d.columns = range(d.shape[1])
    if d.shape[1] < min_cols:
        d = d.reindex(columns=range(min_cols))
    return d


@st.cache_data(show_spinner=False)
def get_template_bytes(b64_string: str) -> bytes:
    """서식 템플릿 base64 디코딩 결과를 캐시. 매 rerun마다 반복 디코딩하지 않는다."""
    return base64.b64decode(b64_string)


@st.cache_data(show_spinner=False)
def build_report_bytes(results_tuple: tuple, target_region: str, template_bytes: bytes) -> bytes:
    """지역별 집계 결과(results)를 서식 워크북에 채워 넣고 바이트로 직렬화한다.
    같은 (집계 결과, 지역, 서식)이면 워크북을 다시 만들지 않고 캐시된 파일을 반환한다."""
    results = dict(results_tuple)
    wb = openpyxl.load_workbook(BytesIO(template_bytes))
    ws = wb.active

    for row in ws.iter_rows(min_row=1, max_row=4):
        for cell in row:
            if type(cell).__name__ != 'MergedCell':
                if isinstance(cell.value, str) and "천안" in cell.value:
                    cell.value = cell.value.replace("천안", target_region)

    ws['B5'] = results[('공사', 1)][0]; ws['B6'] = results[('공사', 1)][1]
    ws['C5'] = results[('공사', 2)][0]; ws['C6'] = results[('공사', 2)][1]
    ws['D5'] = results[('공사', 3)][0]; ws['D6'] = results[('공사', 3)][1]

    ws['F5'] = results[('용역', 1)][0]; ws['F6'] = results[('용역', 1)][1]
    ws['G5'] = results[('용역', 2)][0]; ws['G6'] = results[('용역', 2)][1]
    ws['H5'] = results[('용역', 3)][0]; ws['H6'] = results[('용역', 3)][1]

    ws['J5'] = results[('물품', 1)][0]; ws['J6'] = results[('물품', 1)][1]
    ws['K5'] = results[('물품', 2)][0]; ws['K6'] = results[('물품', 2)][1]
    ws['L5'] = results[('물품', 3)][0]; ws['L6'] = results[('물품', 3)][1]

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


# ---------------------------------------------------------
# 3. 메인 UI 및 데이터 흐름
# ---------------------------------------------------------
st.markdown("""
<div class="brand-header">
    <div class="brand-mark">📊</div>
    <div>
        <div class="brand-title">지역경제활성화 자동 집계 시스템</div>
        <div class="brand-sub">관내구매 실적 집계 · 거래처 주소 자동 채움 · 조달청 나라장터 연동</div>
    </div>
</div>
""", unsafe_allow_html=True)

# [사이드바 0단계] 접속 현황 + 나가기
with st.sidebar:
    _room = get_room_state()
    st.caption(f"🟢 현재 사용 중 {len(_room['active'])}/{MAX_CONCURRENT}명 · 대기 {len(_room['waiting'])}명")
    if st.button("🚪 나가기 (자리 반납)", width='stretch'):
        _room["active"].pop(st.session_state.session_id, None)
        st.success("자리를 반납했습니다. 페이지를 닫으셔도 됩니다.")
        st.stop()
    st.markdown("---")

# [사이드바 1단계] 데이터 업로드
with st.sidebar:
    st.header("📂 데이터 업로드")
    target_region = st.selectbox("기준 지역 선택", CHUNGNAM_REGIONS, index=0)
    data_files = st.file_uploader(
        "자료관리목록 엑셀 업로드",
        type=["xls", "xlsx"],
        accept_multiple_files=True,
    )
    st.markdown(
        '<div class="upload-note">여러 개 업로드 시 자동으로 합쳐서 집계됩니다. '
        '박스 아래 <b>+</b> 를 눌러 파일을 추가하세요.<br>'
        '* 자료관리목록은 학교회계-계약관리-계약자료관리-자료관리에서 '
        '내려받으시기 바랍니다.</div>',
        unsafe_allow_html=True,
    )

    if data_files:
        st.caption(f"📎 현재 {len(data_files)}개 파일 업로드됨 (합산 집계)")

    st.markdown("---")
    st.header("💰 집계 기준")
    amt_col1, amt_col2 = st.columns([3, 1])
    with amt_col1:
        amount_input_raw = st.text_input(
            "금액 기준",
            value=f"{DEFAULT_TARGET_AMOUNT:,}",
            label_visibility="collapsed",
        )
    with amt_col2:
        st.markdown('<div style="padding-top:8px;">원 이상</div>', unsafe_allow_html=True)
    st.caption("숫자만 입력 · 빈칸으로 두면 0원 이상(전체) 집계")
    amount_cleaned = amount_input_raw.replace(",", "").replace(" ", "").strip()
    if amount_cleaned == "":
        TARGET_AMOUNT = 0
    elif amount_cleaned.isdigit():
        TARGET_AMOUNT = int(amount_cleaned)
    else:
        st.warning(f"숫자만 입력해 주세요. 기본값({DEFAULT_TARGET_AMOUNT:,}원)을 사용합니다.")
        TARGET_AMOUNT = DEFAULT_TARGET_AMOUNT

if data_files:
    fingerprint = tuple((f.name, f.size) for f in data_files)
    if st.session_state.get('uploaded_fingerprint') != fingerprint:
        # 업로드된 파일 목록이 바뀐 경우(추가/삭제)에만 다시 합쳐서 읽는다.
        # (그렇지 않으면 화면 재실행마다 주소 수동입력 등 기존 작업 내용이 날아감)
        #
        # ★ 중요: pd.concat은 기본적으로 '컬럼 이름'으로 정렬한다. 이 앱은
        # AMT_COL/BIZ_COL/ADDR_COL처럼 '컬럼 위치(순번)'로 데이터를 읽으므로,
        # 파일마다 헤더 텍스트가 미세하게 달라도(공백 한 칸 등) 컬럼이 밀릴 수
        # 있다. 이를 막기 위해 각 파일의 컬럼명을 0,1,2... 위치 번호로 강제
        # 통일한 뒤(위치 기준 정렬) 합치고, 컬럼 수가 모자란 파일은 필요한
        # 최소 폭(ADDR_COL+1)까지 빈 컬럼으로 채워 IndexError를 방지한다.
        MIN_COLS = 21  # ADDR_COL(20)까지 접근하려면 최소 21개 컬럼 필요
        dfs = [load_excel_normalized(f.getvalue(), MIN_COLS) for f in data_files]
        st.session_state.df = pd.concat(dfs, ignore_index=True)
        st.session_state.uploaded_fingerprint = fingerprint

    df = st.session_state.df
    AMT_COL = 6
    COMP_COL = 16
    BIZ_COL = 18
    ADDR_COL = 20

    biz_col_data = df.iloc[:, BIZ_COL].astype(str).str.strip()
    is_valid_biz = (biz_col_data.str.len() >= 8) & (~biz_col_data.str.contains("Unnamed|nan|None", case=False, na=False))

    amt_data = pd.to_numeric(df.iloc[:, AMT_COL], errors='coerce').fillna(0)
    is_target_amt = amt_data >= TARGET_AMOUNT

    target_mask = is_valid_biz & is_target_amt
    total_target_rows = target_mask.sum()

    missing_mask = df.iloc[:, ADDR_COL].isna() | (df.iloc[:, ADDR_COL].astype(str).str.strip() == "") | (df.iloc[:, ADDR_COL].astype(str).str.strip() == "nan")
    missing_indices = df[missing_mask & target_mask].index
    missing_count = len(missing_indices)

    # --- 메인 화면: KPI 카드 ---
    completion_rate = 0 if total_target_rows == 0 else round((total_target_rows - missing_count) / total_target_rows * 100, 1)
    delta_cls = "ok" if missing_count == 0 else "warn"
    delta_txt = "모두 채워짐" if missing_count == 0 else f"{missing_count}건 조치 필요"

    st.markdown(f"""
    <div class="kpi-row">
        <div class="kpi-card">
            <div class="kpi-label">업로드된 유효 데이터</div>
            <div class="kpi-value">{is_valid_biz.sum():,}<span style="font-size:0.9rem;font-weight:500;color:var(--text-2)">건</span></div>
        </div>
        <div class="kpi-card">
            <div class="kpi-label">집계 대상 ({TARGET_AMOUNT:,}원 이상)</div>
            <div class="kpi-value">{total_target_rows:,}<span style="font-size:0.9rem;font-weight:500;color:var(--text-2)">건</span></div>
        </div>
        <div class="kpi-card">
            <div class="kpi-label">주소 완료율</div>
            <div class="kpi-value">{completion_rate}<span style="font-size:0.9rem;font-weight:500;color:var(--text-2)">%</span></div>
            <div class="kpi-delta {delta_cls}">{delta_txt}</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # [사이드바 2단계] 주소 자동 채우기 버튼 동적 생성
    with st.sidebar:
        st.markdown("---")
        st.header("📍 주소 자동 채우기")
        if st.button("1️⃣ 파일 내 주소 자체 채우기", width='stretch'):
            known_addrs = df.dropna(subset=[df.columns[ADDR_COL]]).set_index(df.columns[BIZ_COL])[df.columns[ADDR_COL]].to_dict()
            df[df.columns[ADDR_COL]] = df[df.columns[ADDR_COL]].fillna(df[df.columns[BIZ_COL]].map(known_addrs))
            st.session_state.df = df
            st.rerun()

        if st.button("2️⃣ 조달청 API 검색", width='stretch'):
            if missing_count == 0:
                st.info("채울 주소가 없습니다.")
            else:
                _room = get_room_state()
                progress_bar = st.progress(0, text="조달청 서버 검색 중...")
                filled_api = 0
                for i, idx in enumerate(missing_indices):
                    addr = get_addr_api(df.iat[idx, BIZ_COL])
                    if addr:
                        df.iat[idx, ADDR_COL] = addr
                        filled_api += 1
                    # 오래 걸리는 작업 중에는 대기시간 만료로 자리를 뺏기지
                    # 않도록 주기적으로 활동시각을 갱신한다.
                    _room["active"][st.session_state.session_id] = time.time()
                    progress_bar.progress((i + 1) / len(missing_indices), text=f"{i+1}/{len(missing_indices)}건 완료")
                st.session_state.df = df
                st.rerun()

    # --- 메인 화면: 수동 입력 표 ---
    if missing_count > 0:
        missing_df = df.loc[missing_indices]
        unique_missing = missing_df.drop_duplicates(subset=[df.columns[BIZ_COL]])

        st.markdown(f"""
        <div class="status-pill warn">⚠ 아직 주소를 찾지 못한 고유 업체가 {len(unique_missing)}곳 있습니다</div>
        """, unsafe_allow_html=True)
        st.write("")

        edit_data = {
            '업체명': unique_missing.iloc[:, COMP_COL].astype(str).replace('nan', '정보없음'),
            '사업자번호': unique_missing.iloc[:, BIZ_COL].astype(str),
            '비즈노 검색링크': unique_missing.iloc[:, BIZ_COL].apply(
                lambda x: f"https://bizno.net/?area=&query={str(x).strip()}" if pd.notna(x) else ""
            ),
            '주소 수동 입력칸': [""] * len(unique_missing)
        }
        edit_df = pd.DataFrame(edit_data, index=unique_missing.index)

        edited_df = st.data_editor(
            edit_df,
            column_config={
                "비즈노 검색링크": st.column_config.LinkColumn("비즈노 링크", display_text="🔍 비즈노에서 찾기"),
                "주소 수동 입력칸": st.column_config.TextColumn("📝 주소 수동 입력칸 (여기에 타이핑)")
            },
            disabled=["업체명", "사업자번호", "비즈노 검색링크"],
            width='stretch'
        )

        if st.button("💾 입력한 주소 원본에 일괄 적용하기"):
            apply_count = 0
            for idx in edited_df.index:
                new_addr = edited_df.at[idx, '주소 수동 입력칸']
                biz_num = edited_df.at[idx, '사업자번호']

                if new_addr and str(new_addr).strip():
                    mask = (df.iloc[:, BIZ_COL].astype(str).str.strip() == biz_num)
                    df.loc[mask, df.columns[ADDR_COL]] = str(new_addr).strip()
                    apply_count += 1

            if apply_count > 0:
                st.session_state.df = df
                st.rerun()
            else:
                st.warning("입력된 주소가 없습니다.")
    else:
        st.markdown(f'<div class="status-pill ok">✓ 집계 대상({TARGET_AMOUNT:,}원 이상)의 주소가 모두 채워졌습니다</div>', unsafe_allow_html=True)
        st.caption("좌측 메뉴에서 결과를 다운로드하세요.")

    # [사이드바 3단계] 결과물 다운로드 및 개발자 정보
    with st.sidebar:
        st.markdown("---")
        st.header("📥 결과물 다운로드")

        raw_output = BytesIO()
        with pd.ExcelWriter(raw_output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
        raw_output.seek(0)

        st.download_button(
            label="📥 1. 주소 완료 원본(Raw)",
            data=raw_output,
            file_name="주소완료_자료관리목록.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width='stretch'
        )

        calc_df = df.copy()
        calc_df['금액'] = pd.to_numeric(calc_df.iloc[:, AMT_COL], errors='coerce')
        calc_df = calc_df[calc_df['금액'] >= TARGET_AMOUNT].copy()

        def categorize_region(addr):
            addr_str = str(addr)
            if target_region in addr_str: return 1
            elif "충남" in addr_str or "충청남도" in addr_str: return 2
            else: return 3

        calc_df['지역구분'] = calc_df.iloc[:, ADDR_COL].apply(categorize_region)

        results = {
            ('공사', 1): [0, 0], ('공사', 2): [0, 0], ('공사', 3): [0, 0],
            ('용역', 1): [0, 0], ('용역', 2): [0, 0], ('용역', 3): [0, 0],
            ('물품', 1): [0, 0], ('물품', 2): [0, 0], ('물품', 3): [0, 0],
        }

        for _, row in calc_df.iterrows():
            g_type = str(row.iloc[1]).strip()
            g_loc = row['지역구분']
            amount = row['금액']
            if (g_type, g_loc) in results:
                results[(g_type, g_loc)][0] += 1
                results[(g_type, g_loc)][1] += amount

        st.session_state['results'] = results
        st.session_state['target_region_used'] = target_region

        try:
            template_bytes = get_template_bytes(TEMPLATE_BASE64)
            # results의 값(list)은 해싱이 안 되므로 tuple로 바꿔 캐시 키로 사용
            results_tuple = tuple((k, tuple(v)) for k, v in results.items())
            report_bytes = build_report_bytes(results_tuple, target_region, template_bytes)

            st.download_button(
                label="🚀 2. 최종 실적보고서(서식)",
                data=report_bytes,
                file_name=f"지역경제활성화_실적보고({target_region}기준).xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width='stretch'
            )
        except Exception as e:
            st.error(f"⚠️ 템플릿 로딩 중 오류가 발생했습니다: {e}")

        st.markdown("---")
        st.caption("개발: 천안버들유치원 나대현<br>문의 및 오류 신고는 메신저로 부탁드립니다.", unsafe_allow_html=True)

    # --- 메인 화면 하단: 지역 분포 바 (시그니처 요소) ---
    results = st.session_state.get('results')
    if results:
        total_amt = sum(v[1] for v in results.values())
        if total_amt > 0:
            amt_1 = sum(results[(t, 1)][1] for t in ['공사', '용역', '물품'])
            amt_2 = sum(results[(t, 2)][1] for t in ['공사', '용역', '물품'])
            amt_3 = sum(results[(t, 3)][1] for t in ['공사', '용역', '물품'])
            pct_1 = round(amt_1 / total_amt * 100, 1)
            pct_2 = round(amt_2 / total_amt * 100, 1)
            pct_3 = round(amt_3 / total_amt * 100, 1)
            region_label = st.session_state.get('target_region_used', target_region)

            st.markdown(f"""
            <div class="region-bar-card">
                <div class="region-bar-title">💠 지역별 구매 금액 비중 ({region_label} 기준)</div>
                <div class="region-bar-track">
                    <div class="region-bar-seg" style="width:{pct_1}%; background:var(--blue);"></div>
                    <div class="region-bar-seg" style="width:{pct_2}%; background:var(--mint);"></div>
                    <div class="region-bar-seg" style="width:{pct_3}%; background:#CBD3E1;"></div>
                </div>
                <div class="region-legend">
                    <span><span class="dot" style="background:var(--blue);"></span>{region_label} {pct_1}%</span>
                    <span><span class="dot" style="background:var(--mint);"></span>충남(관내 외) {pct_2}%</span>
                    <span><span class="dot" style="background:#CBD3E1;"></span>기타 {pct_3}%</span>
                </div>
            </div>
            """, unsafe_allow_html=True)

else:
    st.markdown("""
    <div class="kpi-card" style="text-align:center; padding: 40px 20px; color:var(--text-2);">
        👈 왼쪽 사이드바에서 자료관리목록 엑셀 파일을 먼저 업로드해 주세요.
    </div>
    """, unsafe_allow_html=True)

    with st.sidebar:
        st.markdown("---")
        st.caption("개발: 천안버들유치원 나대현<br>문의 및 오류 신고는 메신저로 부탁드립니다.", unsafe_allow_html=True)
