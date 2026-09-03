from io import BytesIO

import streamlit as st
from openpyxl import load_workbook

from core_logic import categorize_region_code, format_money, safe_value
import excel_reports as reports
from template_fallback import load_halfyear_template_safe

reports._load_halfyear_template = lambda: load_halfyear_template_safe(reports.HALFYEAR_TEMPLATE)

OFFICIAL_GUIDE = "◑에듀파인-계약목록 내려받기 후(①2025년회계2026.1월~2월,② 2026년회계3월~7월)→목록 정리(공공요금, 수수료 등삭제)→ 엑셀.net 활용(파일 올리기)"


def _restore_official_review_header(file_bytes):
    wb = load_workbook(BytesIO(file_bytes))
    ws = wb[reports.REVIEW_SHEET]
    ws['A4'] = OFFICIAL_GUIDE
    ws['J4'] = '지역경제활성화 자동 집계 시스템 · Streamlit'
    out = BytesIO(); wb.save(out)
    return out.getvalue()


def _short_half(label):
    return '상' if str(label).startswith('상') else ('하' if str(label).startswith('하') else str(label))


def render_mode1_outputs(ctx):
    df = st.session_state.df
    report_mask = ctx['report_mask']
    contract_types = ctx['contract_types']
    col_addr = ctx['col_addr']
    col_amount = ctx['col_amount']
    target_region = ctx['target_region']

    results = {(t, loc): [0, 0] for t in ['공사', '용역', '물품'] for loc in [1, 2, 3]}
    for idx in df.index[report_mask]:
        row = df.loc[idx]
        ctype = contract_types.at[idx]
        loc = categorize_region_code(safe_value(row, col_addr, ''), target_region)
        amount = format_money(safe_value(row, col_amount, 0))
        results[ctype, loc][0] += 1
        results[ctype, loc][1] += amount

    total = sum(v[1] for v in results.values())
    if total:
        amounts = [sum(results[t, loc][1] for t in ['공사', '용역', '물품']) for loc in [1, 2, 3]]
        pct = [round(x / total * 100, 1) for x in amounts]
        st.markdown(
            f'''<div class="region-card compact-region"><div class="work-title">지역별 구매 금액 비중</div><div class="region-track"><div class="region-seg-1" style="width:{pct[0]}%"></div><div class="region-seg-2" style="width:{pct[1]}%"></div><div class="region-seg-3" style="width:{pct[2]}%"></div></div><div class="region-legend"><span>{target_region} {pct[0]}%</span><span>충남 관외 {pct[1]}%</span><span>타시도 {pct[2]}%</span></div></div>''',
            unsafe_allow_html=True,
        )

    st.markdown('<div class="section-title compact-title">결과 파일</div>', unsafe_allow_html=True)
    col1, col2 = st.columns(2, gap='small')

    with col1:
        with st.container(border=True):
            h1, h2 = st.columns([0.18, 0.82], gap='small')
            with h1:
                st.markdown('<div class="small-file-icon">▤</div>', unsafe_allow_html=True)
            with h2:
                st.markdown('<div class="card-title">분기별 실적보고서</div><div class="card-desc no-margin">현재 사용 중인 분기 제출서식</div>', unsafe_allow_html=True)
            try:
                q = reports.build_quarter_report_bytes(results, target_region)
                st.download_button(
                    '분기보고서 다운로드',
                    q,
                    f'지역경제활성화_실적보고({target_region}기준).xlsx',
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    width='stretch',
                    key='quarter_download',
                )
            except Exception as exc:
                st.error(f'분기보고서 생성 오류: {exc}')

    with col2:
        with st.container(border=True):
            h1, h2 = st.columns([0.18, 0.82], gap='small')
            with h1:
                st.markdown('<div class="small-file-icon green-file">✓</div>', unsafe_allow_html=True)
            with h2:
                st.markdown('<div class="card-title">반기 검토용 기초자료</div><div class="card-desc no-margin">1-4 양식 · 검토 후 최종작성에서 재업로드</div>', unsafe_allow_html=True)
            try:
                b, n, _ = reports.build_review_workbook_bytes(
                    df,
                    ctx['headers'],
                    ctx['target_amount'],
                    target_region,
                    '',
                    '',
                    ctx['report_year'],
                    ctx['report_label'],
                    ctx['period_start'],
                    ctx['period_end'],
                )
                b = _restore_official_review_header(b)
                half = _short_half(ctx['report_label'])
                st.download_button(
                    f'검토용 파일 다운로드 · {n:,}건',
                    b,
                    f'지역경제활성화_{ctx["report_year"]}{half}_검토용.xlsx',
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    width='stretch',
                    key='review_download',
                )
            except Exception as exc:
                st.error(f'검토용 파일 생성 오류: {exc}')
