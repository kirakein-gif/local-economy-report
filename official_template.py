import base64
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
PART_DIR = BASE_DIR / "template_parts"
REQUIRED_SHEETS = [
    "1-1. 총괄(공사)",
    "1-2. 총괄(용역)",
    "1-3. 총괄(물품)",
    "1-4. 기초자료(지역경제활성화)",
]
PART_SPECS = [
    ("official_01.b64", 8000),
    ("official_02.b64", 8000),
    ("official_03a.b64", 4000),
    ("official_03b.b64", 4000),
    ("official_04.b64", 4948),
]


def load_official_halfyear_template():
    """사용자가 제공한 확정 양식을 보존한 내장 템플릿을 복원합니다."""
    chunks = []
    for name, expected_len in PART_SPECS:
        path = PART_DIR / name
        if not path.exists():
            raise FileNotFoundError(f"공식 반기양식 내장 데이터가 없습니다: {name}")
        text = path.read_text(encoding="utf-8").strip()
        if len(text) < expected_len:
            raise RuntimeError(f"공식 반기양식 데이터가 불완전합니다: {name}")
        # 저장 커넥터에서 줄 끝 문자가 붙는 경우에도 원본 길이만 정확히 사용합니다.
        chunks.append(text[:expected_len])

    try:
        raw = base64.b64decode("".join(chunks), validate=True)
        if not raw.startswith(b"PK\x03\x04"):
            raise ValueError("XLSX ZIP 헤더가 올바르지 않습니다.")
        wb = load_workbook(BytesIO(raw))
    except Exception as exc:
        raise RuntimeError(f"공식 반기양식을 복원하지 못했습니다: {exc}") from exc

    if not set(REQUIRED_SHEETS).issubset(set(wb.sheetnames)):
        raise RuntimeError("공식 반기양식의 4개 시트를 확인하지 못했습니다.")

    ws1 = wb["1-1. 총괄(공사)"]
    ws3 = wb["1-3. 총괄(물품)"]
    ws4 = wb["1-4. 기초자료(지역경제활성화)"]
    if (
        ws1["A2"].value != "순"
        or ws1["F2"].value != "총괄"
        or ws3["N2"].value != "구입목적별 "
        or "<작성방법>" not in str(ws4["A2"].value or "")
        or ws4["J5"].value != "금액단위: 원"
    ):
        raise RuntimeError("복원된 반기양식이 확정 양식 구조와 일치하지 않습니다.")

    return wb
