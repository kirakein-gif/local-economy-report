"""Reporting labels only: never filter or reclassify contract records."""
from collections import Counter
from datetime import date, datetime
from io import BytesIO
import re

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

DATE_PART = r'(\d{4})\s*[.\-/년]\s*(\d{1,2})\s*[.\-/월]\s*(\d{1,2})\s*일?\.?'
PERIOD = re.compile(DATE_PART + r'\s*[~～–]\s*' + DATE_PART)
YEAR_HALF = re.compile(r'(\d{4})년\s*(상반기|하반기|반기)')


def contract_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or '').strip()
    try:
        if re.fullmatch(r'\d{8}', text):
            return datetime.strptime(text, '%Y%m%d').date()
        match = re.match(DATE_PART, text)
        if match:
            return date(*map(int, match.groups()))
        if isinstance(value, (int, float)) and 1 <= value <= 73050:
            parsed = from_excel(value)
            return parsed.date() if isinstance(parsed, datetime) else None
    except (ValueError, OverflowError):
        pass
    return None


def infer_report_info(records, title, today=None):
    today = today or date.today()
    dates = [d for row in records if (d := contract_date(row.get('계약일자')))]
    title_match = YEAR_HALF.search(title or '')
    title_year = int(title_match[1]) if title_match else None
    title_half = title_match[2] if title_match and title_match[2] != '반기' else None
    counts = Counter(d.year for d in dates)
    warnings = []
    if counts:
        candidates = sorted(y for y, n in counts.items() if n == max(counts.values()))
        year = title_year if title_year in candidates else candidates[-1]
        if len(candidates) > 1:
            warnings.append('보고연도별 계약 건수가 같습니다. 보고연도를 확인해 주세요.')
    else:
        year = title_year or today.year
        warnings.append('계약일을 인식하지 못했습니다. 보고연도와 기간을 확인해 주세요.')
    first = sum(d.year == year and d.month <= 6 for d in dates)
    second = sum(d.year == year and d.month >= 7 for d in dates)
    if first == second:
        half = title_half or '상반기'
        warnings.append('상·하반기 건수가 같거나 없습니다. 보고구분을 확인해 주세요.')
    else:
        half = '상반기' if first > second else '하반기'
    period = PERIOD.search(title or '')
    start = end = None
    if period:
        try:
            start = date(*map(int, period.groups()[:3]))
            end = date(*map(int, period.groups()[3:]))
            if start > end:
                start = end = None
        except ValueError:
            start = end = None
    source = '검토파일 제목의 집계기간'
    if start is None:
        source = '계약일의 최소·최대 날짜'
        start, end = (min(dates), max(dates)) if dates else (date(year, 1, 1), date(year, 12, 31))
        if not dates:
            source = '임시 기본값 · 직접 확인 필요'
    if len(dates) != len(records):
        warnings.append(f'계약일 미인식 {len(records) - len(dates):,}건은 자동 추천에서 제외했습니다.')
    if len(counts) > 1:
        warnings.append('여러 연도 자료가 포함되어 있습니다. 모든 계약은 그대로 집계합니다.')
    return dict(year=year, half=half, start=start, end=end, first=first, second=second,
                source=source, warnings=warnings)


def report_filename(year, half, institution):
    name = re.sub(r'[/\\:*?"<>|\x00-\x1f]', '_', str(institution or '').strip()) or '기관명미입력'
    return f'지역경제활성화 실적 작성 양식({year}{"상" if half == "상반기" else "하"}_{name}).xlsx'


def apply_report_info(content, year, half, start, end):
    """Change A1 report headings/periods; preserve cells, formulas and aggregation."""
    if not 1900 <= year <= 2100 or half not in ('상반기', '하반기') or start > end:
        raise ValueError('보고연도·구분·기간을 확인해 주세요.')
    wb = load_workbook(BytesIO(content))
    period = f'{start:%Y. %m. %d.} ~ {end:%Y. %m. %d.}'
    for ws in wb.worksheets:
        title = str(ws['A1'].value or '')
        if YEAR_HALF.search(title):
            title = YEAR_HALF.sub(f'{year}년 {half}', title, count=1)
        else:
            title = re.sub(r'(지역경제활성화)', f'{year}년 {half} 지역경제활성화', title, count=1)
        if PERIOD.search(title):
            title = PERIOD.sub(period, title, count=1)
        else:
            title += f' [집계기간: {period}]'
        ws['A1'] = title
    result = BytesIO()
    wb.save(result)
    return result.getvalue()
