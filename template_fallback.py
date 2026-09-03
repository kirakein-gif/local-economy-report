from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SUMMARY_SHEETS = ["1-1. 총괄(공사)", "1-2. 총괄(용역)", "1-3. 총괄(물품)"]
REVIEW_SHEET = "1-4. 기초자료(지역경제활성화)"

WHITE = "FFFFFF"
BLACK = "000000"
HEADER_FILL = "D9EAF7"
SUB_FILL = "EAF3F8"
LINE = "7F7F7F"

OFFICIAL_GUIDE = "◑에듀파인-계약목록 내려받기 후(①2025년회계2026.1월~2월,② 2026년회계3월~7월)→목록 정리(공공요금, 수수료 등삭제)→ 엑셀.net 활용(파일 올리기)"
OFFICIAL_METHOD = """<작성방법>
1. 지역: 소속 기관(학교)이 위치한 시·군명을 입력 (예: 홍성, 천안, 공주 등)
2. 기관명(학교명): 소속 기관 또는 학교명 전체 입력 (예: 홍성초등학교)
3.급별: 학교(유), 학교(초), 학교(중), 학교(고), 학교(특수), 교육지원청, 직속기관 중 선택
4. 계약방법: 수의계약/입찰/조달구매 중 선택
5. 견적/경쟁방법: 1인수의(단일견적)/2인수의(공개견적)/제3자단가/다수공급자2단계경쟁(MAS)/우수조달/중앙조달/자체(제한경쟁)/자체(일반경쟁)/협상/2단계(규격가격동시)
6. 주소: 도&시군까지만 기재(예시: 충남 홍성군, 충남 천안시, 서울특별시 영등포구, 경기도 수원시..)
6. 소재지 (★ 중요): 계약 상대자(업체)의 주소지 기준으로 선택
-충남도내(관내): 소속 기관과 동일한 시·군 소재 업체
-충남도내(관외): 소속 기관 외 충남 도내 타 시·군 소재 업체
-타시도: 충청남도 외 타 시·도 소재 업체
7. 구입목적: 교육용[늘(돌)봄포함 하되 간식제외), 도서(교육용간행물 포함 하되 신문 등 업무용 간행물 및 교과용도서 구입은 제외함)], 그 외 중 선택
※주의사항: 조달청 종합쇼핑몰, 학교장터(S2B) 등의 사이트 이용 시, 계약상대자(업체)의 실제 사업자등록증 상 주소를 기준으로 입력"""


def _font(size=10, bold=False):
    return Font(name="맑은 고딕", size=size, bold=bold, color=BLACK)


def _border():
    s = Side(style="thin", color=LINE)
    return Border(left=s, right=s, top=s, bottom=s)


def _cell(cell, fill=None, bold=False, size=10, horizontal="center"):
    cell.font = _font(size, bold)
    cell.alignment = Alignment(horizontal=horizontal, vertical="center", wrap_text=True)
    cell.border = _border()
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)


def _merge(ws, refs):
    for ref in refs:
        ws.merge_cells(ref)


def _summary_common(ws, title, purpose):
    _merge(ws, ["A1:M1", "A2:A4", "B2:B4", "C2:C4", "D2:D4", "E2:E4", "F2:M2", "F3:G3", "H3:I3", "J3:K3", "L3:M3"])
    ws["A1"] = title
    ws["A2"] = "순"; ws["B2"] = "지역"; ws["C2"] = "기관명(학교명)"; ws["D2"] = "급별"; ws["E2"] = "목적물별"; ws["F2"] = "총괄"
    ws["F3"] = "충남도내(관내)"; ws["H3"] = "충남도내(관외)"; ws["J3"] = "충남도외(타시도)"; ws["L3"] = "합계"
    for c, text in zip(range(6, 14), ["건수", "금액", "건수", "금액", "건수", "금액", "건수", "금액"]):
        ws.cell(4, c).value = text
    ws["A5"] = "(예시)"; ws["B5"] = "천안 "; ws["C5"] = "**초등학교"; ws["D5"] = "학교(초)"; ws["E5"] = purpose
    for r in range(2, 6):
        for c in range(1, 14):
            if type(ws.cell(r, c)).__name__ != "MergedCell":
                _cell(ws.cell(r, c), HEADER_FILL if r <= 4 else WHITE, r <= 4)
    ws["A1"].font = _font(14, True); ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 22; ws.row_dimensions[4].height = 22; ws.row_dimensions[5].height = 24
    widths = [7, 12, 22, 12, 12] + [11] * 8
    for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_view.showGridLines = False


def _setup_work(ws):
    _summary_common(ws, "붙임 1-1. 2026년 상반기 지역경제활성화 추진 실적(총괄+공사)", "공사")
    ws.merge_cells("A7:J7"); ws["A7"] = OFFICIAL_GUIDE; ws["K7"] = "지역경제활성화 자동 집계 시스템 · Streamlit"


def _setup_service(ws):
    _summary_common(ws, "붙임 1-2. 2026년 상반기 지역경제활성화 추진 실적(총괄+용역)", "용역")


def _setup_goods(ws):
    _merge(ws, ["A1:M1", "A2:A5", "B2:B5", "C2:C5", "D2:D5", "E2:E5", "F2:M3", "N2:AC2", "N3:U3", "V3:AC3", "F4:G4", "H4:I4", "J4:K4", "L4:M4", "N4:O4", "P4:Q4", "R4:S4", "T4:U4", "V4:W4", "X4:Y4", "Z4:AA4", "AB4:AC4"])
    ws["A1"] = "붙임 1-3. 2026년 상반기 지역경제활성화 추진 실적(총괄+물품)"
    ws["A2"] = "순"; ws["B2"] = "지역"; ws["C2"] = "기관명(학교명)"; ws["D2"] = "급별"; ws["E2"] = "목적물별"; ws["F2"] = "총괄"; ws["N2"] = "구입목적별 "
    ws["N3"] = "교육용 물품 구입"; ws["V3"] = "도서구입"
    groups = [("F4", "충남도내(관내)"), ("H4", "충남도내(관외)"), ("J4", "충남도외(타시도)"), ("L4", "합계"), ("N4", "충청남도내(관내)"), ("P4", "충청남내(관외)"), ("R4", "타시도"), ("T4", "합계"), ("V4", "충청남도내(관내)"), ("X4", "충청남내(관외)"), ("Z4", "타시도"), ("AB4", "합계")]
    for ref, text in groups: ws[ref] = text
    for c in range(6, 30):
        ws.cell(5, c).value = "건수" if c % 2 == 0 else "금액"
    ws["A6"] = "(예시)"; ws["B6"] = "천안 "; ws["C6"] = "**초등학교"; ws["D6"] = "학교(초)"; ws["E6"] = "물품"
    for r in range(2, 7):
        for c in range(1, 30):
            if type(ws.cell(r, c)).__name__ != "MergedCell": _cell(ws.cell(r, c), HEADER_FILL if r <= 5 else WHITE, r <= 5, 9)
    ws["A1"].font = _font(14, True); ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    for i, w in enumerate([7, 12, 22, 12, 12] + [10] * 24, 1): ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    for r in range(2, 7): ws.row_dimensions[r].height = 23
    ws.sheet_view.showGridLines = False


def _setup_review(ws):
    _merge(ws, ["A1:O1", "A2:O2", "A4:I4"])
    ws["A1"] = "붙임1-4. 2026년 상반기 지역경제활성화 추진 실적[기준(건당 10만원 이상), 2026. 1. 1. ~ 2026. 7. 31.]"
    ws["A2"] = OFFICIAL_METHOD
    ws["A4"] = OFFICIAL_GUIDE
    ws["J4"] = "지역경제활성화 자동 집계 시스템 · Streamlit"
    ws["J5"] = "금액단위: 원"
    headers = ["순", "지역", "기관명\n(학교명)", "급별", "목적물", "계약방법", "견적/경쟁방법", "계약명", "계약일자", "계약금액(원)", "업체명", "주소", "소재지", "구입목적", "비고"]
    for c, text in enumerate(headers, 1):
        ws.cell(6, c).value = text; _cell(ws.cell(6, c), HEADER_FILL, True, 9)
    example = ["(예시)", "천안 ", "**초등학교", "", "", "", "1인수의(단일견적)", "", "", 0, "", "충청남도 천안시", "충남도내(관내)", "", ""]
    for c, value in enumerate(example, 1):
        ws.cell(7, c).value = value; _cell(ws.cell(7, c), WHITE, False, 9)
    widths = [7, 10, 22, 12, 10, 14, 20, 42, 13, 14, 22, 34, 17, 18, 24]
    for c, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 28; ws.row_dimensions[2].height = 138; ws.row_dimensions[4].height = 28; ws.row_dimensions[6].height = 38; ws.row_dimensions[7].height = 24
    ws.sheet_view.showGridLines = False


def build_fallback_workbook():
    wb = Workbook()
    ws1 = wb.active; ws1.title = SUMMARY_SHEETS[0]; _setup_work(ws1)
    ws2 = wb.create_sheet(SUMMARY_SHEETS[1]); _setup_service(ws2)
    ws3 = wb.create_sheet(SUMMARY_SHEETS[2]); _setup_goods(ws3)
    ws4 = wb.create_sheet(REVIEW_SHEET); _setup_review(ws4)
    return wb


def _looks_official(wb):
    try:
        return (
            set(SUMMARY_SHEETS + [REVIEW_SHEET]).issubset(set(wb.sheetnames))
            and wb[SUMMARY_SHEETS[0]]["A2"].value == "순"
            and wb[SUMMARY_SHEETS[0]]["F2"].value == "총괄"
            and wb[SUMMARY_SHEETS[2]]["N2"].value == "구입목적별 "
            and "<작성방법>" in str(wb[REVIEW_SHEET]["A2"].value or "")
            and wb[REVIEW_SHEET]["J5"].value == "금액단위: 원"
        )
    except Exception:
        return False


def load_halfyear_template_safe(template_path):
    """공식 템플릿 구조까지 검증하고, 손상/변형 시 공식 구조의 내장 양식으로 복구합니다."""
    try:
        p = Path(template_path)
        if p.exists():
            wb = load_workbook(p)
            if _looks_official(wb):
                return wb
    except Exception:
        pass
    return build_fallback_workbook()
