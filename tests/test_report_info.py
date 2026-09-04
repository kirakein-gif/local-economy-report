import unittest
from datetime import date
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

import excel_reports as reports
from official_template import load_official_halfyear_template
from report_info import apply_report_info, contract_date, infer_report_info, report_filename


class ReportInfoTests(unittest.TestCase):
    def test_january_to_july_uses_majority_without_truncating_period(self):
        rows = [{'계약일자': d} for d in ('2026-02-01', '2026-06-30', '2026-07-31')]
        info = infer_report_info(rows, '2026년 상반기 [2026. 1. 1. ~ 2026. 7. 31.]')
        self.assertEqual((info['year'], info['half'], info['first'], info['second']),
                         (2026, '상반기', 2, 1))
        self.assertEqual((info['start'], info['end']), (date(2026, 1, 1), date(2026, 7, 31)))
        self.assertFalse(info['warnings'])

    def test_dates_ties_missing_and_multiple_years(self):
        for value in ('20260731', '2026. 7. 31.', '2026-07-31', 46234):
            self.assertEqual(contract_date(value), date(2026, 7, 31))
        self.assertIsNone(contract_date('2026-02-31'))
        self.assertTrue(infer_report_info([{'계약일자': ''}], '')['warnings'])
        tied = [{'계약일자': '2026-01-01'}, {'계약일자': '2026-07-31'}]
        self.assertTrue(infer_report_info(tied, '')['warnings'])
        years = [{'계약일자': '2026-01-01'}, {'계약일자': '2027-01-01'}]
        self.assertTrue(infer_report_info(years, '')['warnings'])
        self.assertEqual(infer_report_info([{'계약일자': '2026-08-01'}], '2026년 상반기')['half'], '하반기')

    def test_real_report_only_changes_headings(self):
        reports._load_halfyear_template = load_official_halfyear_template
        headers = [f'열{i}' for i in range(21)]
        for index, label in {1: '계약구분', 2: '계약방법', 4: '계약명', 5: '계약일자',
                             6: '계약금액', 16: '업체명', 18: '사업자등록번호', 20: '주소'}.items():
            headers[index] = label
        rows = []
        for kind, day in [('공사', '2026-01-01'), ('용역', '2026-06-30'), ('물품', '2026-07-31')]:
            row = [''] * 21
            for index, value in {1: kind, 2: '수의계약', 4: '검증 계약', 5: day, 6: 700000,
                                 16: '검증업체', 18: '1111111111', 20: '충청남도 천안시 테스트로'}.items():
                row[index] = value
            rows.append(row)
        review, _, _ = reports.build_review_workbook_bytes(pd.DataFrame(rows), headers, 0, '천안',
            '검증기관', '학교(유)', 2026, '상반기', date(2026, 1, 1), date(2026, 7, 31))
        original = reports.build_final_halfyear_report_bytes(review)[0]
        updated = apply_report_info(original, 2027, '하반기', date(2027, 7, 1), date(2027, 12, 31))
        before, after = load_workbook(BytesIO(original)), load_workbook(BytesIO(updated))
        self.assertEqual(before.sheetnames, after.sheetnames)
        self.assertEqual(len(after.sheetnames), 4)
        for name in before.sheetnames:
            self.assertIn('2027년 하반기', after[name]['A1'].value)
            self.assertIn('2027. 07. 01.', after[name]['A1'].value)
            self.assertEqual(list(before[name].merged_cells.ranges), list(after[name].merged_cells.ranges))
            for row in before[name]:
                for cell in row:
                    if cell.coordinate != 'A1':
                        self.assertEqual(cell.value, after[name][cell.coordinate].value)
                        self.assertEqual(cell.style_id, after[name][cell.coordinate].style_id)
        with self.assertRaises(ValueError):
            apply_report_info(original, 2026, '상반기', date(2026, 8, 1), date(2026, 7, 1))
        self.assertIn('2027하_검증_기관', report_filename(2027, '하반기', '검증/기관'))


if __name__ == '__main__':
    unittest.main()
